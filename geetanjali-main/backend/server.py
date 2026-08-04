"""Geetanjali Salon — Agentic Middleware Backend (PostgreSQL Edition)
Complete migration from MongoDB to PostgreSQL via SQLAlchemy async.
All API endpoints return identical JSON shapes — frontend requires zero changes.
"""
from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import csv
import uuid
import logging
import bcrypt
import jwt as pyjwt
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests as httpreq
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional, Dict, Any
from collections import defaultdict

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, UploadFile, File, Depends, Query
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from sqlalchemy import select, update, delete, func, distinct, and_, or_, text
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from core.pg_database import engine, async_session, init_pg, close_pg, get_session
from core.pg_models import (
    Base, User, Staff, SKU, SKUBatch, POSTransaction, POSTransactionStaff,
    Payout, Checkout, PurchaseInvoice, PurchaseInvoiceLine,
    Vendor, AppConfig, AppFlag,
    Attendance, StockAudit, StockAuditItem,
    ServiceRecipe, RecipeIngredient, ServiceConsumptionLog,
    Budget, BudgetLineItem,
    PurchaseOrder, POLine, POStatusHistory,
    VendorContract, StockLedger, ProductIncentiveMapping,
)

from core.incentive_engine import (
    DEFAULT_CONFIG, get_config, calc_eligible_service_amount, calc_staff_eligible_value,
    calc_daily_bonus, calc_monthly_bonus, calc_manager_bonus, calc_product_incentive
)

# ------------------ Config ------------------
JWT_SECRET = os.environ.get("JWT_SECRET", "supersecret")
JWT_ALG = "HS256"

app = FastAPI(title="Geetanjali Salon Platform")

api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("lss")

# ------------------ Helpers ------------------
def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()

def new_id() -> str:
    return str(uuid.uuid4())

def hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def verify_pw(pw: str, hashed: str) -> bool:
    return bcrypt.checkpw(pw.encode(), hashed.encode())

def make_token(user_id: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "access",
    }
    return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except pyjwt.PyJWTError:
        raise HTTPException(401, "Invalid or expired token")
    async with async_session() as session:
        result = await session.execute(select(User).where(User.id == payload["sub"]))
        user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(401, "User not found")
    return user.to_dict()

def require_role(*roles):
    async def _dep(user: dict = Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(403, f"Requires role: {roles}")
        return user
    return _dep

def parse_date_flex(v) -> Optional[str]:
    """Return YYYY-MM-DD or None."""
    if not v or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

def to_float(v) -> float:
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return 0.0
        s = str(v).strip().replace(",", "")
        if s == "" or s.lower() == "nan":
            return 0.0
        return float(s)
    except (ValueError, TypeError):
        return 0.0

# ------------------ Default Incentive Config ------------------
DEFAULT_CONFIG = {
    "id": "master",
    "staff_daily_tiers": [
        {"min": 2500, "max": 4999, "bonus": 100},
        {"min": 5000, "max": 7999, "bonus": 200},
        {"min": 8000, "max": 9999, "bonus": 350},
        {"min": 10000, "max": 14999, "bonus": 500},
        {"min": 15000, "max": 17999, "bonus": 700},
        {"min": 18000, "max": 99999999, "bonus": 1000},
    ],
    "video_review_bonus": 50,
    "staff_monthly_multipliers": [
        {"min_ratio": 4, "max_ratio": 5, "pct": 3},
        {"min_ratio": 5, "max_ratio": 6, "pct": 5},
        {"min_ratio": 6, "max_ratio": 9999, "pct": 6},
    ],
    "retail_commission_pct": 0,
    "product_incentives": [
        # L'Oreal Professionnel
        {"brand": "loreal", "brand_display": "L'Oréal Professionnel", "min_price": 1, "max_price": 2000, "amount": 50},
        {"brand": "loreal", "brand_display": "L'Oréal Professionnel", "min_price": 2001, "max_price": 4000, "amount": 100},
        {"brand": "loreal", "brand_display": "L'Oréal Professionnel", "min_price": 4001, "max_price": 999999, "amount": 150},
        # Kanpeki
        {"brand": "kanpeki", "brand_display": "Kanpeki", "min_price": 1, "max_price": 3000, "amount": 50},
        {"brand": "kanpeki", "brand_display": "Kanpeki", "min_price": 3001, "max_price": 999999, "amount": 100},
        # Kerastase
        {"brand": "kerastase", "brand_display": "Kérastase", "min_price": 1, "max_price": 3000, "amount": 50},
        {"brand": "kerastase", "brand_display": "Kérastase", "min_price": 3001, "max_price": 6000, "amount": 100},
        {"brand": "kerastase", "brand_display": "Kérastase", "min_price": 6001, "max_price": 9000, "amount": 150},
        {"brand": "kerastase", "brand_display": "Kérastase", "min_price": 9001, "max_price": 12000, "amount": 200},
        {"brand": "kerastase", "brand_display": "Kérastase", "min_price": 12001, "max_price": 999999, "amount": 250},
        # Olaplex
        {"brand": "olaplex", "brand_display": "Olaplex", "min_price": 1, "max_price": 3000, "amount": 50},
        {"brand": "olaplex", "brand_display": "Olaplex", "min_price": 3001, "max_price": 5000, "amount": 100},
        {"brand": "olaplex", "brand_display": "Olaplex", "min_price": 5001, "max_price": 999999, "amount": 150},
        # Kerafusion / De Fabulous
        {"brand": "de fabulous", "brand_display": "Kerafusion / De Fabulous", "min_price": 1, "max_price": 2000, "amount": 50},
        {"brand": "de fabulous", "brand_display": "Kerafusion / De Fabulous", "min_price": 2001, "max_price": 4000, "amount": 100},
        {"brand": "de fabulous", "brand_display": "Kerafusion / De Fabulous", "min_price": 4001, "max_price": 999999, "amount": 150},
        # Amazon Series
        {"brand": "amazon series", "brand_display": "Amazon Series", "min_price": 1, "max_price": 999999, "amount": 100},
        # QOD
        {"brand": "qod", "brand_display": "QOD", "min_price": 1, "max_price": 999999, "amount": 100},
    ],
    "manager_milestones": [
        {"min_revenue": 1800000, "bonus_per_manager": 5000},
        {"min_revenue": 2000000, "bonus_per_manager": 7000},
        {"min_revenue": 2500000, "bonus_per_manager": 10000},
        {"min_revenue": 3000000, "bonus_per_manager": 20000},
    ],
    "inventory": {
        "lead_time_days": 4,
        "safety_buffer_pct": 50,
    },
    "prepaid_card_bonuses": [
        {"pattern": "11,000", "amount": 300},
        {"pattern": "21,000", "amount": 500},
        {"pattern": "51,000", "amount": 1000},
        {"pattern": "infinity", "amount": 2000},
    ],
}

# ------------------ Models (Pydantic — request schemas) ------------------
class LoginIn(BaseModel):
    email: str
    password: str

class CheckoutIn(BaseModel):
    sku_id: str
    quantity: float
    notes: Optional[str] = ""

class ReceiveIn(BaseModel):
    sku_id: str
    quantity: float
    unit_cost: float = 0
    expiry_date: Optional[str] = None

class PurchaseInvoiceLineIn(BaseModel):
    sku_id: str
    quantity: float
    unit_cost: float
    expiry_date: Optional[str] = None

class PurchaseInvoiceIn(BaseModel):
    invoice_number: str
    vendor: str
    invoice_date: str
    lines: List[PurchaseInvoiceLineIn]
    notes: Optional[str] = ""

class SKUCreateIn(BaseModel):
    name: str
    category: str = "Uncategorized"
    unit_cost: float = 0
    unit_price: float = 0
    opening_store_qty: float = 0
    opening_floor_qty: float = 0
    opening_expiry: Optional[str] = None

class ConfirmPayoutIn(BaseModel):
    staff_id: str
    payout_date: str
    amount: float
    breakdown: Dict[str, Any]

class ConfigUpdateIn(BaseModel):
    staff_daily_tiers: Optional[List[Dict[str, Any]]] = None
    video_review_bonus: Optional[float] = None
    staff_monthly_multipliers: Optional[List[Dict[str, Any]]] = None
    retail_commission_pct: Optional[float] = None
    manager_milestones: Optional[List[Dict[str, Any]]] = None
    inventory: Optional[Dict[str, Any]] = None
    prepaid_card_bonuses: Optional[List[Dict[str, Any]]] = None
    product_incentives: Optional[List[Dict[str, Any]]] = None

class ProductMappingIn(BaseModel):
    pos_item_name: str
    brand: Optional[str] = ""
    pattern: Optional[str] = ""
    sku_id: Optional[str] = None
    amount: float
    min_price: Optional[float] = None
    max_price: Optional[float] = None

class VendorIn(BaseModel):
    name: str
    lead_time_days: int = 4
    contact: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    address: Optional[str] = ""
    gst_number: Optional[str] = ""
    notes: Optional[str] = ""

# ------------------ Incentive Engine ------------------
async def get_config() -> dict:
    async with async_session() as session:
        result = await session.execute(select(AppConfig).where(AppConfig.id == "master"))
        cfg = result.scalar_one_or_none()
    if cfg and cfg.data:
        d = dict(cfg.data)
        d["id"] = "master"
        return d
    return DEFAULT_CONFIG

def calc_daily_bonus(service_revenue: float, tiers: List[dict]) -> Dict[str, Any]:
    tier_hit = None
    total = float(service_revenue or 0.0)
    for t in sorted(tiers, key=lambda x: x["min"], reverse=True):
        if total >= t["min"]:
            tier_hit = t
            break
    return {
        "service_revenue": round(total, 2),
        "tier": tier_hit,
        "bonus": tier_hit["bonus"] if tier_hit else 0,
    }

def calc_monthly_bonus(monthly_service_rev: float, salary: float, mults: List[dict]) -> Dict[str, Any]:
    if salary <= 0:
        return {"ratio": 0, "pct": 0, "amount": 0}
    ratio = monthly_service_rev / salary
    hit = None
    for m in sorted(mults, key=lambda x: x["min_ratio"], reverse=True):
        if ratio >= m["min_ratio"]:
            hit = m
            break
    return {
        "ratio": round(ratio, 2),
        "pct": hit["pct"] if hit else 0,
        "amount": round(monthly_service_rev * (hit["pct"] / 100), 2) if hit else 0,
    }

def calc_manager_bonus(month_revenue: float, milestones: List[dict]) -> Dict[str, Any]:
    hit = None
    for m in sorted(milestones, key=lambda x: x["min_revenue"]):
        if month_revenue >= m["min_revenue"]:
            hit = m
    return {
        "month_revenue": round(month_revenue, 2),
        "milestone": hit,
        "bonus": hit["bonus_per_manager"] if hit else 0,
    }

def calc_product_incentive(item_name: str, brand: str, net_price: float, qty: float, rules: List[dict], mappings: Optional[Dict[str, dict]] = None) -> float:
    if not item_name or qty <= 0:
        return 0.0
    item_key = item_name.strip().lower()

    # 1. Check custom saved product mappings first (exact POS product name match)
    if mappings and item_key in mappings:
        map_entry = mappings[item_key]
        return float(map_entry.get("amount", 0)) * qty

    name_lc = item_key
    brand_lc = (brand or "").lower()
    aliases = {
        "loreal": ["l'oreal", "loreal", "l'oreal professionnel", "loreal professionnel", "serie expert", "absolut repair", "metal dx", "aminexil", "serioxyl", "vitamino", "inoa", "majirel"],
        "kanpeki": ["kanpeki", "kenpeki"],
        "kerastase": ["kerastase", "kérastase", "k chroma", "k genesis", "k reflection", "k nutritive", "k specifique", "k densifique", "k blond", "k discipline", "k resistance", "k elixir", "k initialiste", "k symbiose", "k first", " keras", "elixir ultime"],
        "olaplex": ["olaplex", "bond maintenance"],
        "de fabulous": ["de fabulous", "kerafusion"],
        "amazon series": ["amazon", "amazone", "amazon series", "amazone series"],
        "qod": ["qod"],
    }
    unit_price = net_price / qty if qty else 0
    for rule in rules:
        rb = (rule.get("brand") or "").lower()
        if rb:
            hit = rb in name_lc or rb in brand_lc
            if not hit:
                for a in aliases.get(rb, []):
                    if a in name_lc:
                        hit = True
                        break
            if not hit:
                continue
        pat = (rule.get("pattern") or "").lower()
        if pat:
            pat_aliases = {"shampoo": ["shmp", "shampooing"], "masque": ["mask", "masq"],
                            "conditioner": ["cond"], "serum": ["srm"]}
            hit_pat = pat in name_lc
            if not hit_pat:
                for a in pat_aliases.get(pat, []):
                    if a in name_lc:
                        hit_pat = True
                        break
            if not hit_pat:
                continue
        mn = rule.get("min_price")
        mx = rule.get("max_price")
        if mn is not None and unit_price < mn:
            continue
        if mx is not None and unit_price > mx:
            continue
        return float(rule.get("amount", 0)) * qty
    return 0.0

async def _staff_day_product_incentive(staff_name: str, day: str, rules: List[dict], mappings: Optional[Dict[str, dict]] = None) -> float:
    total = 0.0
    staff_lc = staff_name.strip().lower()
    async with async_session() as session:
        if mappings is None:
            map_rows = (await session.execute(select(ProductIncentiveMapping))).scalars().all()
            mappings = {m.pos_item_name: m.to_dict() for m in map_rows}

        q = (
            select(POSTransaction)
            .options(selectinload(POSTransaction.staff_shares))
            .join(POSTransactionStaff)
            .where(
                POSTransaction.date == day,
                func.lower(func.trim(POSTransaction.type)) == "product",
                func.lower(func.trim(POSTransactionStaff.name)) == staff_lc,
            )
        )
        result = await session.execute(q)
        rows = result.scalars().unique().all()
        for r in rows:
            share = next((s.pct for s in r.staff_shares if s.name.strip().lower() == staff_lc), 100) / 100
            brand = ""
            sku_result = await session.execute(select(SKU).where(SKU.name == (r.item_name or "").strip()))
            sku = sku_result.scalar_one_or_none()
            if sku:
                brand = sku.vendor_name or sku.category or ""
            total += calc_product_incentive(r.item_name or "", brand, r.net_price or 0, r.quantity or 1, rules, mappings) * share
    return round(total, 2)
    return round(total, 2)


# ------------------ POS Import ------------------
def parse_pos_row(row: dict, fallback_date: Optional[str] = None, row_idx: int = 0) -> Optional[dict]:
    item_name = str(row.get("Item Name") or "").strip()
    if not item_name or item_name.lower() in ("nan", "none", "total", "subtotal", "grand total", "summary"):
        return None

    date_str = (
        parse_date_flex(row.get("Date"))
        or parse_date_flex(row.get("Invoice Date"))
        or parse_date_flex(row.get("Bill Date"))
        or parse_date_flex(row.get("Date & Time"))
        or parse_date_flex(row.get("Txn Date"))
        or fallback_date
    )
    if not date_str:
        return None

    inv_no = str(row.get("Invoice Number") or "").strip()
    if not inv_no or inv_no.lower() in ("nan", "none"):
        inv_no = f"INV-ROW-{row_idx + 1}"

    net_price = to_float(row.get("Net Price"))
    rate = to_float(row.get("Rate"))
    qty = to_float(row.get("Quantity")) or 1
    total_discount = (
        to_float(row.get("Membership Discount"))
        + to_float(row.get("Manager Discount"))
        + to_float(row.get("Offer Discount"))
    )
    gross = rate * qty
    is_full_discount = gross > 0 and net_price == 0 and total_discount > 0

    vc_val = None
    for k in ["Paid from Value Card", "Paid from VC", "Value Card Paid", "Value Card", "VC Paid", "VC", "Paid from ValueCard"]:
        if k in row and row[k] is not None and str(row[k]).strip().lower() != "nan":
            vc_val = to_float(row[k])
            break
    if vc_val is None:
        vc_val = to_float(row.get("Other"))
    vc_paid = max(0.0, float(vc_val or 0.0))

    staff = []
    for i in range(1, 5):
        name = row.get(f"Staff {i}")
        if name and str(name).strip() and str(name).strip().lower() != "nan":
            pct = to_float(row.get(f"Staff {i} %")) or 100
            staff.append({"name": str(name).strip(), "pct": pct})

    client_name = str(
        row.get("Client")
        or row.get("Client Name")
        or row.get("Customer")
        or row.get("Customer Name")
        or row.get("Party Name")
        or ""
    ).strip()

    return {
        "id": new_id(),
        "salon": str(row.get("Salon") or "").strip(),
        "invoice_number": inv_no,
        "date": date_str,
        "time": str(row.get("Time") or "").strip(),
        "client": client_name,
        "type": str(row.get("Type") or "Service").strip(),
        "item_name": item_name,
        "category": str(row.get("Category") or "").strip(),
        "quantity": qty,
        "rate": rate,
        "membership_discount": to_float(row.get("Membership Discount")),
        "manager_discount": to_float(row.get("Manager Discount")),
        "offer_discount": to_float(row.get("Offer Discount")),
        "total_discount": total_discount,
        "net_price": net_price,
        "tax": to_float(row.get("Tax")),
        "total_collection": to_float(row.get("Total Collection")),
        "cash": to_float(row.get("Cash")),
        "card": to_float(row.get("Card")),
        "other": vc_paid,
        "staff": staff,
        "is_quality_failure": is_full_discount and str(row.get("Type") or "").strip().lower() == "service",
        "created_at": now_utc(),
    }

async def import_csv_bytes(csv_bytes: bytes) -> Dict[str, Any]:
    df = pd.DataFrame()
    if csv_bytes.startswith(b'PK') or csv_bytes.startswith(b'\xd0\xcf') or csv_bytes.startswith(b'\x09\x08'):
        try:
            df_raw = pd.read_excel(io.BytesIO(csv_bytes), header=None)
            header_idx = 0
            for idx, row in df_raw.iterrows():
                row_str = " ".join([str(val).lower() for val in row.values if pd.notna(val)])
                if ("invoice" in row_str or "bill" in row_str or "client" in row_str or "voucher" in row_str or "doc" in row_str) and ("date" in row_str or "item" in row_str or "net" in row_str or "qty" in row_str):
                    header_idx = idx
                    break
            df = pd.read_excel(io.BytesIO(csv_bytes), skiprows=header_idx)
        except Exception as ex:
            log.error(f"Excel read error: {ex}")
            try:
                df = pd.read_excel(io.BytesIO(csv_bytes))
            except Exception as ex2:
                log.error(f"Excel fallback read error: {ex2}")
                df = pd.DataFrame()
    else:
        text = csv_bytes.decode("utf-8", errors="replace")
        lines = text.splitlines()
        header_idx = 0
        for i, ln in enumerate(lines[:30]):
            ln_lc = ln.lower()
            # Must have multiple commas (actual CSV columns) and contain key column names
            comma_count = ln.count(",")
            has_key_cols = (
                ("invoice number" in ln_lc or "invoice no" in ln_lc or "bill no" in ln_lc)
                or ("item name" in ln_lc or "item_name" in ln_lc)
                or ("net price" in ln_lc or "net amount" in ln_lc)
            )
            if comma_count >= 3 and has_key_cols:
                header_idx = i
                break
        log.info(f"CSV header row detected at line {header_idx}: {lines[header_idx][:80] if lines else 'N/A'}")
        trimmed = "\n".join(lines[header_idx:])
        try:
            df = pd.read_csv(io.StringIO(trimmed), on_bad_lines="skip")
        except Exception:
            try:
                df = pd.read_csv(io.StringIO(text), on_bad_lines="skip")
            except Exception as ex:
                log.error(f"CSV read error: {ex}")
                df = pd.DataFrame()

    canonical_map = {
        "salon": "Salon",
        "invoice number": "Invoice Number", "invoice no": "Invoice Number", "invoice": "Invoice Number", "bill no": "Invoice Number", "bill number": "Invoice Number", "voucher no": "Invoice Number", "document no": "Invoice Number",
        "date": "Date", "txn date": "Date", "transaction date": "Date", "invoice date": "Date", "bill date": "Date", "sales date": "Date", "sale date": "Date", "date & time": "Date", "date/time": "Date", "datetime": "Date", "posting date": "Date", "created date": "Date",
        "time": "Time", "txn time": "Time",
        "client": "Client", "client name": "Client", "customer": "Client", "customer name": "Client", "patient": "Client", "party name": "Client",
        "phone": "Phone", "mobile": "Phone", "contact": "Phone",
        "type": "Type", "item type": "Type",
        "item name": "Item Name", "service name": "Item Name", "product name": "Item Name", "item": "Item Name", "particulars": "Item Name",
        "category": "Category", "item category": "Category",
        "hsn code": "HSN Code", "hsn": "HSN Code",
        "quantity": "Quantity", "qty": "Quantity",
        "rate": "Rate", "unit rate": "Rate", "price": "Rate",
        "membership discount": "Membership Discount", "manager discount": "Manager Discount", "offer discount": "Offer Discount",
        "net price": "Net Price", "net amount": "Net Price", "net value": "Net Price", "net": "Net Price", "amount": "Net Price",
        "paid from service balance": "Paid from Service Balance",
        "paid from value card": "Paid from Value Card", "paid from vc": "Paid from Value Card", "value card paid": "Paid from Value Card", "value card": "Paid from Value Card", "vc paid": "Paid from Value Card", "vc": "Paid from Value Card", "paid from valuecard": "Paid from Value Card",
        "taxable price": "Taxable Price", "taxable amount": "Taxable Price", "taxable": "Taxable Price",
        "tax": "Tax", "gst": "Tax", "tax amount": "Tax", "tax %": "Tax %",
        "change to advance": "Change to Advance",
        "total collection": "Total Collection", "total amount": "Total Collection", "gross total": "Total Collection", "total": "Total Collection",
        "cash": "Cash", "card": "Card", "other": "Other",
        "paid from advance": "Paid from Advance", "paid from gift card": "Paid from Gift Card",
        "staff 1": "Staff 1", "staff 1 %": "Staff 1 %", "staff 2": "Staff 2", "staff 2 %": "Staff 2 %",
        "staff 3": "Staff 3", "staff 3 %": "Staff 3 %", "staff 4": "Staff 4", "staff 4 %": "Staff 4 %",
    }
    new_cols = []
    for c in df.columns:
        key = str(c).strip().lower()
        new_cols.append(canonical_map.get(key, str(c).strip()))
    df.columns = new_cols

    records = []
    last_date = datetime.now().strftime("%Y-%m-%d")
    for row_idx, (_, r) in enumerate(df.iterrows()):
        parsed = parse_pos_row(r.to_dict(), fallback_date=last_date, row_idx=row_idx)
        if parsed:
            if parsed.get("date"):
                last_date = parsed["date"]
            records.append(parsed)
    if not records:
        return {"imported": 0, "quality_failures": 0}

    async with async_session() as session:
        new_records = []
        staff_names = set()

        log.info(f"import_csv_bytes: {len(records)} parsed records to process")

        # Collect all unique invoice numbers being uploaded
        invoice_numbers = list({rec["invoice_number"] for rec in records})
        log.info(f"import_csv_bytes: {len(invoice_numbers)} unique invoice numbers")

        # Delete existing transactions for these invoices (allows clean reimport)
        if invoice_numbers:
            existing_ids_result = await session.execute(
                select(POSTransaction.id).where(POSTransaction.invoice_number.in_(invoice_numbers))
            )
            existing_ids = [r[0] for r in existing_ids_result.all()]
            if existing_ids:
                await session.execute(
                    delete(POSTransactionStaff).where(POSTransactionStaff.transaction_id.in_(existing_ids))
                )
                await session.execute(
                    delete(POSTransaction).where(POSTransaction.id.in_(existing_ids))
                )
                log.info(f"import_csv_bytes: deleted {len(existing_ids)} existing rows for reimport")

        # Insert all records fresh
        for rec in records:
            staff_list = rec.pop("staff", [])
            try:
                session.add(POSTransaction(**rec))
                new_records.append(rec)
                for s in staff_list:
                    cname = str(s.get("name") or "").strip()
                    if cname:
                        staff_names.add(cname)
                        session.add(POSTransactionStaff(transaction_id=rec["id"], name=cname, pct=s.get("pct", 100)))
            except Exception as ex:
                log.error(f"Failed to insert row {rec.get('invoice_number')}: {ex}")

        # Seed staff from POS
        for name in staff_names:
            existing = await session.execute(
                select(Staff).where(func.lower(func.trim(Staff.name)) == name.lower())
            )
            if not existing.scalar_one_or_none():
                session.add(Staff(id=new_id(), name=name, base_salary=25000, role="staff", created_at=now_utc()))

        await session.commit()
        log.info(f"import_csv_bytes committed: {len(new_records)} rows inserted fresh")


        # Seed SKUs from products and auto-checkout
        auto_co = 0
        async with async_session() as session2:
            for rec in new_records:
                if rec["type"].lower() == "product" and rec["item_name"]:
                    sku_name = rec["item_name"].strip()
                    existing = await session2.execute(select(SKU).where(SKU.name == sku_name))
                    sku = existing.scalar_one_or_none()
                    if not sku:
                        sku = SKU(
                            id=new_id(), name=sku_name,
                            category=rec["category"] or "Uncategorized",
                            unit_cost=round(rec["rate"] * 0.55, 2),
                            unit_price=rec["rate"],
                            store_qty=12, floor_qty=3, retail_qty=0,
                            created_at=now_utc(),
                        )
                        session2.add(sku)
                        # Add default batches
                        session2.add(SKUBatch(id=new_id(), sku_id=sku.id, qty=3, location="floor",
                                             unit_cost=round(rec["rate"] * 0.55, 2), received_at=now_utc()))
                        session2.add(SKUBatch(id=new_id(), sku_id=sku.id, qty=12, location="store",
                                             unit_cost=round(rec["rate"] * 0.55, 2), received_at=now_utc()))
                        await session2.flush()

                    # Auto-checkout from retail/floor for POS product sale
                    consumed = await _consume_batches_session(session2, sku.id, "retail", rec["quantity"])
                    if consumed > 0:
                        session2.add(Checkout(
                            id=new_id(), sku_id=sku.id, sku_name=sku.name,
                            quantity=consumed, notes=f"POS retail sale · invoice {rec['invoice_number']}",
                            source="pos", invoice_number=rec["invoice_number"],
                            checked_out_by="system", checked_out_at=now_utc(),
                        ))
                        auto_co += 1
            await session2.commit()

        qf = sum(1 for r in new_records if r.get("is_quality_failure"))
        return {"imported": len(new_records), "new_rows": len(new_records), "updated_rows": 0, "quality_failures": qf, "pos_auto_checkouts": auto_co}


# ------------------ Batch helpers ------------------
async def _consume_batches_session(session: AsyncSession, sku_id: str, location: str, qty_needed: float) -> float:
    """FIFO by expiry_date (None last). Returns qty actually consumed. Uses existing session."""
    result = await session.execute(
        select(SKUBatch).where(SKUBatch.sku_id == sku_id, SKUBatch.location == location, SKUBatch.qty > 0)
        .order_by(func.coalesce(SKUBatch.expiry_date, '9999-12-31'), SKUBatch.received_at)
    )
    batches = list(result.scalars().all())
    remaining = qty_needed
    for b in batches:
        if remaining <= 0:
            break
        take = min(b.qty, remaining)
        b.qty -= take
        remaining -= take
    consumed = qty_needed - remaining
    if consumed <= 0:
        return 0
    # Clean up empty batches
    for b in batches:
        if b.qty <= 0:
            await session.delete(b)
    # Recompute totals
    await _recompute_sku_totals(session, sku_id)
    return consumed


async def _move_batches_session(session: AsyncSession, sku_id: str, from_loc: str, to_loc: str, qty: float) -> float:
    """Move qty FIFO from one location to another. Uses existing session."""
    result = await session.execute(
        select(SKUBatch).where(SKUBatch.sku_id == sku_id, SKUBatch.location == from_loc, SKUBatch.qty > 0)
        .order_by(func.coalesce(SKUBatch.expiry_date, '9999-12-31'), SKUBatch.received_at)
    )
    src_batches = list(result.scalars().all())
    remaining = qty
    for b in src_batches:
        if remaining <= 0:
            break
        take = min(b.qty, remaining)
        b.qty -= take
        session.add(SKUBatch(
            id=new_id(), sku_id=sku_id, qty=take, location=to_loc,
            expiry_date=b.expiry_date, unit_cost=b.unit_cost,
            invoice_id=b.invoice_id, received_at=now_utc(),
        ))
        remaining -= take
    moved = qty - remaining
    if moved <= 0:
        return 0
    for b in src_batches:
        if b.qty <= 0:
            await session.delete(b)
    await _recompute_sku_totals(session, sku_id)
    return moved


async def _recompute_sku_totals(session: AsyncSession, sku_id: str):
    """Recalculate store_qty, floor_qty, retail_qty from batches."""
    result = await session.execute(
        select(SKUBatch.location, func.sum(SKUBatch.qty))
        .where(SKUBatch.sku_id == sku_id, SKUBatch.qty > 0)
        .group_by(SKUBatch.location)
    )
    totals = {row[0]: row[1] for row in result.all()}
    await session.execute(
        update(SKU).where(SKU.id == sku_id).values(
            store_qty=totals.get("store", 0),
            floor_qty=totals.get("floor", 0),
            retail_qty=totals.get("retail", 0),
        )
    )


# ------------------ AUTH ------------------
@api.post("/auth/login")
async def login(payload: LoginIn, response: Response):
    email = payload.email.lower().strip()
    async with async_session() as session:
        result = await session.execute(select(User).where(User.email == email))
        u = result.scalar_one_or_none()
    log.info(f"Login attempt: email={email}, user_found={bool(u)}")
    if not u or not verify_pw(payload.password, u.password_hash):
        raise HTTPException(401, "Invalid credentials")
    token = make_token(u.id, u.role)
    response.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=604800, path="/")
    return {"id": u.id, "email": u.email, "name": u.name, "role": u.role, "token": token}

@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

# ------------------ POS ------------------
@api.post("/pos/upload")
async def pos_upload(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    contents = await file.read()
    result = await import_csv_bytes(contents)
    return result

@api.get("/pos/transactions")
async def list_transactions(date: Optional[str] = None, staff: Optional[str] = None, limit: int = 500):
    async with async_session() as session:
        q = select(POSTransaction).options(selectinload(POSTransaction.staff_shares))
        if date:
            q = q.where(POSTransaction.date == date)
        if staff:
            q = q.join(POSTransactionStaff).where(POSTransactionStaff.name == staff)
        q = q.order_by(POSTransaction.date.desc()).limit(limit)
        result = await session.execute(q)
        return [t.to_dict() for t in result.scalars().unique().all()]

@api.get("/pos/dates")
async def list_pos_dates():
    async with async_session() as session:
        result = await session.execute(select(distinct(POSTransaction.date)))
        dates = [r[0] for r in result.all()]
    return sorted(dates, reverse=True)

@api.post("/pos/reset")
async def pos_reset(user: dict = Depends(get_current_user)):
    async with async_session() as session:
        await session.execute(delete(POSTransaction))
        await session.commit()
    return {"message": "All POS transaction data cleared successfully."}


@api.get("/pos/quality-failures")
async def quality_failures():
    async with async_session() as session:
        q = select(POSTransaction).options(selectinload(POSTransaction.staff_shares)).where(
            POSTransaction.is_quality_failure == True
        ).order_by(POSTransaction.date.desc()).limit(500)
        result = await session.execute(q)
        return [t.to_dict() for t in result.scalars().unique().all()]

async def _auto_sync_staff_from_pos():
    """Auto-seed any staff member present in POS transaction shares into Staff table with Rudrapur July salary structure."""
    salary_defaults = [
        ("suhail khan", (50000, "Manager", "manager")),
        ("khan", (50000, "Manager", "manager")),
        ("uhd suhail", (41000, "Stylist", "stylist")),
        ("suhail", (41000, "Stylist", "stylist")),
        ("siraj", (43000, "Stylist", "stylist")),
        ("jahangir", (35000, "Stylist", "stylist")),
        ("ashu", (70000, "Stylist", "stylist")),
        ("faid", (20000, "Assistant", "assistant")),
        ("sadik", (33000, "Barber", "barber")),
        ("alam", (23000, "Barber", "barber")),
        ("anju", (21000, "Beautician", "beautician")),
        ("urosha", (35000, "Beautician", "beautician")),
        ("uroosha", (35000, "Beautician", "beautician")),
        ("navneet", (27000, "Beautician", "beautician")),
        ("soni", (15000, "Housekeeping", "housekeeping")),
        ("lalita", (15000, "Housekeeping", "housekeeping")),
        ("geeta", (13000, "Housekeeping", "housekeeping")),
        ("fahim", (35000, "Pedicurist", "pedicurist")),
        ("faheem", (35000, "Pedicurist", "pedicurist")),
        ("sameer", (25000, "Pedicurist", "pedicurist")),
        ("sandhaya", (43000, "Manager", "manager")),
        ("sandhya", (43000, "Manager", "manager")),
    ]
    async with async_session() as session:
        distinct_names = (await session.execute(
            select(POSTransactionStaff.name).where(
                POSTransactionStaff.name.isnot(None), POSTransactionStaff.name != ""
            ).distinct()
        )).scalars().all()
        added = False
        for name in distinct_names:
            cname = name.strip()
            if not cname:
                continue
            if "abhimanyu" in cname.lower() or "dhingra" in cname.lower():
                continue
            existing = (await session.execute(
                select(Staff).where(func.lower(func.trim(Staff.name)) == cname.lower())
            )).scalar_one_or_none()

            # Determine default salary and designation
            c_sal, c_desig, c_role = 25000, "Staff", "staff"
            for key, val in salary_defaults:
                if key in cname.lower():
                    c_sal, c_desig, c_role = val
                    break

            if not existing:
                session.add(Staff(id=new_id(), name=cname, base_salary=c_sal, role=c_role, department=c_desig, created_at=now_utc()))
                added = True
            else:
                # Upgrade existing base_salary if default 25000
                if existing.base_salary != c_sal:
                    existing.base_salary = c_sal
                    existing.department = c_desig
                    existing.role = c_role
                    added = True

        if added:
            await session.commit()


# ------------------ Staff ------------------
class StaffCreateIn(BaseModel):
    name: str
    role: str = "staff"
    base_salary: float = 25000.0
    email: Optional[str] = ""
    phone: Optional[str] = ""
    department: Optional[str] = ""

class StaffUpdateIn(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    base_salary: Optional[float] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    department: Optional[str] = None

@api.get("/staff")
async def list_staff():
    await _auto_sync_staff_from_pos()
    async with async_session() as session:
        result = await session.execute(select(Staff).order_by(Staff.name))
        return [s.to_dict() for s in result.scalars().all()]

@api.post("/staff")
async def create_staff(payload: StaffCreateIn, user: dict = Depends(require_role("owner", "manager", "admin"))):
    async with async_session() as session:
        existing = (await session.execute(
            select(Staff).where(func.lower(func.trim(Staff.name)) == payload.name.strip().lower())
        )).scalar_one_or_none()
        if existing:
            raise HTTPException(400, "Staff member with this name already exists")
        
        new_staff = Staff(
            id=new_id(),
            name=payload.name.strip(),
            email=(payload.email or "").strip(),
            phone=(payload.phone or "").strip(),
            base_salary=payload.base_salary,
            role=payload.role.strip(),
            department=(payload.department or "").strip(),
            created_at=now_utc()
        )
        session.add(new_staff)
        await session.commit()
        return new_staff.to_dict()

@api.put("/staff/{staff_id}")
async def update_staff(staff_id: str, payload: StaffUpdateIn, user: dict = Depends(require_role("owner", "manager", "admin"))):
    async with async_session() as session:
        result = await session.execute(select(Staff).where(Staff.id == staff_id))
        staff = result.scalar_one_or_none()
        if not staff:
            raise HTTPException(404, "Staff not found")
        
        if payload.name is not None:
            staff.name = payload.name.strip()
        if payload.role is not None:
            staff.role = payload.role.strip()
        if payload.base_salary is not None:
            staff.base_salary = payload.base_salary
        if payload.email is not None:
            staff.email = payload.email.strip()
        if payload.phone is not None:
            staff.phone = payload.phone.strip()
        if payload.department is not None:
            staff.department = payload.department.strip()
            
        await session.commit()
        return staff.to_dict()

@api.delete("/staff/{staff_id}")
async def delete_staff(staff_id: str, user: dict = Depends(require_role("owner", "manager", "admin"))):
    async with async_session() as session:
        result = await session.execute(select(Staff).where(Staff.id == staff_id))
        staff = result.scalar_one_or_none()
        if not staff:
            raise HTTPException(404, "Staff not found")
        await session.delete(staff)
        await session.commit()
        return {"success": True}

# ------------------ Incentives ------------------
async def _staff_day_revenue(staff_name: str, day: str) -> Dict[str, float]:
    staff_lc = staff_name.strip().lower()
    async with async_session() as session:
        q = (
            select(
                POSTransaction.type,
                POSTransaction.net_price,
                POSTransaction.other,
                POSTransactionStaff.pct
            )
            .join(POSTransactionStaff)
            .where(
                POSTransaction.date == day,
                func.lower(func.trim(POSTransactionStaff.name)) == staff_lc
            )
        )
        result = await session.execute(q)
        service = 0.0
        retail = 0.0
        for row in result.all():
            t_type, net_price, other_val, share_pct = row[0], row[1] or 0.0, row[2] or 0.0, row[3] or 100.0
            t_type_str = str(t_type).strip().lower()
            if t_type_str == "service":
                eligible_amt = calc_eligible_service_amount(net_price, other_val)
                service += calc_staff_eligible_value(eligible_amt, share_pct)
            elif t_type_str in ("product", "retail"):
                retail += net_price * (share_pct / 100.0)
    return {"service": round(service, 2), "retail": round(retail, 2)}

async def _staff_month_revenue(staff_name: str, month: str) -> Dict[str, float]:
    staff_lc = staff_name.strip().lower()
    async with async_session() as session:
        q = (
            select(
                POSTransaction.type,
                POSTransaction.net_price,
                POSTransaction.other,
                POSTransactionStaff.pct
            )
            .join(POSTransactionStaff)
            .where(
                POSTransaction.date.like(f"{month}%"),
                func.lower(func.trim(POSTransactionStaff.name)) == staff_lc
            )
        )
        result = await session.execute(q)
        service = 0.0
        retail = 0.0
        for row in result.all():
            t_type, net_price, other_val, share_pct = row[0], row[1] or 0.0, row[2] or 0.0, row[3] or 100.0
            t_type_str = str(t_type).strip().lower()
            if t_type_str == "service":
                eligible_amt = calc_eligible_service_amount(net_price, other_val)
                service += calc_staff_eligible_value(eligible_amt, share_pct)
            elif t_type_str in ("product", "retail"):
                retail += net_price * (share_pct / 100.0)
    return {"service": round(service, 2), "retail": round(retail, 2)}

@api.get("/incentives/daily")
async def incentives_daily(day: str):
    await _auto_sync_staff_from_pos()
    cfg = await get_config()
    async with async_session() as session:
        result = await session.execute(
            select(Staff).where(func.lower(Staff.role) != "owner", ~func.lower(Staff.name).contains("abhimanyu"))
        )
        staff_list = result.scalars().all()
    out = []
    for s in staff_list:
        rev = await _staff_day_revenue(s.name, day)
        bonus = calc_daily_bonus(rev["service"], cfg["staff_daily_tiers"])
        product_inc = await _staff_day_product_incentive(s.name, day, cfg.get("product_incentives", []))
        legacy_pct = cfg.get("retail_commission_pct", 0)
        retail_comm = round(rev["retail"] * (legacy_pct / 100), 2) if legacy_pct else 0
        async with async_session() as session2:
            payout_result = await session2.execute(
                select(Payout).where(Payout.staff_id == s.id, Payout.payout_date == day)
            )
            payout = payout_result.scalar_one_or_none()
        out.append({
            "staff_id": s.id, "staff_name": s.name, "base_salary": s.base_salary,
            "service_revenue": bonus["service_revenue"], "retail_revenue": round(rev["retail"], 2),
            "tier": bonus["tier"], "daily_bonus": bonus["bonus"],
            "product_incentive": product_inc, "retail_commission": retail_comm,
            "total_earned": bonus["bonus"] + product_inc + retail_comm,
            "confirmed": bool(payout),
            "confirmed_at": payout.confirmed_at if payout else None,
        })
    return {"day": day, "config": cfg["staff_daily_tiers"], "rows": out}

@api.get("/incentives/daily/details")
async def incentives_daily_details(staff_name: str, day: str):
    staff_lc = staff_name.strip().lower()
    cfg = await get_config()
    rules = cfg.get("product_incentives", [])
    async with async_session() as session:
        s_res = await session.execute(select(Staff).where(func.lower(func.trim(Staff.name)) == staff_lc))
        staff_obj = s_res.scalar_one_or_none()
        base_salary = staff_obj.base_salary if staff_obj else 0.0
        
        map_rows = (await session.execute(select(ProductIncentiveMapping))).scalars().all()
        mappings = {m.pos_item_name: m.to_dict() for m in map_rows}
        q = (
            select(POSTransaction)
            .options(selectinload(POSTransaction.staff_shares))
            .join(POSTransactionStaff)
            .where(
                POSTransaction.date == day,
                func.lower(func.trim(POSTransactionStaff.name)) == staff_lc,
            )
            .order_by(POSTransaction.time)
        )
        result = await session.execute(q)
        rows = result.scalars().unique().all()
        
        details = []
        total_service = 0.0
        total_retail = 0.0
        total_incentive = 0.0
        for r in rows:
            share_row = next((s for s in r.staff_shares if s.name.strip().lower() == staff_lc), None)
            share_pct = share_row.pct if share_row else 100.0
            
            t_type = (r.type or "").strip().lower()
            incentive = 0.0
            if t_type in ("product", "retail"):
                share_value = round((r.net_price or 0) * (share_pct / 100), 2)
                brand = ""
                sku_result = await session.execute(select(SKU).where(SKU.name == (r.item_name or "").strip()))
                sku = sku_result.scalar_one_or_none()
                if sku:
                    brand = sku.vendor_name or sku.category or ""
                incentive = calc_product_incentive(r.item_name or "", brand, r.net_price or 0, r.quantity or 1, rules, mappings) * (share_pct / 100)
                incentive = round(incentive, 2)
                total_retail += share_value
                total_incentive += incentive
            else:
                eligible_svc_amt = calc_eligible_service_amount(r.net_price or 0.0, r.other or 0.0)
                share_value = calc_staff_eligible_value(eligible_svc_amt, share_pct)
                total_service += share_value
            
            details.append({
                "transaction_id": r.id,
                "date": r.date,
                "invoice_number": r.invoice_number,
                "client": r.client,
                "client_name": r.client,
                "item_name": r.item_name,
                "type": r.type,
                "quantity": r.quantity,
                "net_price": r.net_price,
                "value_card_paid": r.other or 0.0,
                "eligible_service_amount": calc_eligible_service_amount(r.net_price or 0.0, r.other or 0.0) if t_type not in ("product", "retail") else r.net_price,
                "share_pct": share_pct,
                "share_value": share_value,
                "incentive": incentive,
            })
            
    daily_bonus = calc_daily_bonus(total_service, cfg["staff_daily_tiers"])
    legacy_pct = cfg.get("retail_commission_pct", 0)
    retail_comm = round(total_retail * (legacy_pct / 100), 2) if legacy_pct else 0
    
    return {
        "staff_name": staff_name,
        "day": day,
        "base_salary": base_salary,
        "total_service": round(total_service, 2),
        "total_retail": round(total_retail, 2),
        "tier": daily_bonus["tier"],
        "daily_bonus": daily_bonus["bonus"],
        "product_incentive": total_incentive,
        "retail_commission": retail_comm,
        "total_earned": daily_bonus["bonus"] + total_incentive + retail_comm,
        "details": details,
        "config_tiers": cfg["staff_daily_tiers"]
    }

@api.get("/incentives/monthly/details")
async def incentives_monthly_details(staff_name: str, month: str):
    staff_lc = staff_name.strip().lower()
    cfg = await get_config()
    rules = cfg.get("product_incentives", [])
    async with async_session() as session:
        s_res = await session.execute(select(Staff).where(func.lower(func.trim(Staff.name)) == staff_lc))
        staff_obj = s_res.scalar_one_or_none()
        base_salary = staff_obj.base_salary if staff_obj else 0.0
        
        map_rows = (await session.execute(select(ProductIncentiveMapping))).scalars().all()
        mappings = {m.pos_item_name: m.to_dict() for m in map_rows}
        q = (
            select(POSTransaction)
            .options(selectinload(POSTransaction.staff_shares))
            .join(POSTransactionStaff)
            .where(
                POSTransaction.date.like(f"{month}%"),
                func.lower(func.trim(POSTransactionStaff.name)) == staff_lc,
            )
            .order_by(POSTransaction.date, POSTransaction.time)
        )
        result = await session.execute(q)
        rows = result.scalars().unique().all()
        
        details = []
        total_service = 0.0
        total_retail = 0.0
        total_incentive = 0.0
        for r in rows:
            share_row = next((s for s in r.staff_shares if s.name.strip().lower() == staff_lc), None)
            share_pct = share_row.pct if share_row else 100.0
            share_value = round((r.net_price or 0) * (share_pct / 100), 2)
            
            t_type = (r.type or "").strip().lower()
            incentive = 0.0
            if t_type in ("product", "retail"):
                share_value = round((r.net_price or 0) * (share_pct / 100), 2)
                brand = ""
                sku_result = await session.execute(select(SKU).where(SKU.name == (r.item_name or "").strip()))
                sku = sku_result.scalar_one_or_none()
                if sku:
                    brand = sku.vendor_name or sku.category or ""
                incentive = calc_product_incentive(r.item_name or "", brand, r.net_price or 0, r.quantity or 1, rules, mappings) * (share_pct / 100)
                incentive = round(incentive, 2)
                total_retail += share_value
                total_incentive += incentive
            else:
                eligible_svc_amt = calc_eligible_service_amount(r.net_price or 0.0, r.other or 0.0)
                share_value = calc_staff_eligible_value(eligible_svc_amt, share_pct)
                total_service += share_value
            
            details.append({
                "transaction_id": r.id,
                "date": r.date,
                "invoice_number": r.invoice_number,
                "client": r.client,
                "client_name": r.client,
                "item_name": r.item_name,
                "type": r.type,
                "quantity": r.quantity,
                "net_price": r.net_price,
                "value_card_paid": r.other or 0.0,
                "eligible_service_amount": calc_eligible_service_amount(r.net_price or 0.0, r.other or 0.0) if t_type not in ("product", "retail") else r.net_price,
                "share_pct": share_pct,
                "share_value": share_value,
                "incentive": incentive,
            })
            
    monthly_bonus = calc_monthly_bonus(total_service, base_salary, cfg["staff_monthly_multipliers"])
    retail_comm_pct = cfg.get("retail_commission_pct", 0)
    retail_comm = round(total_retail * (retail_comm_pct / 100), 2)
    
    # Calculate prepaid card sale bonuses
    prepaid_card_bonus_sum = 0.0
    bonuses = cfg.get("prepaid_card_bonuses", [])
    for r in rows:
        share_row = next((s for s in r.staff_shares if s.name.strip().lower() == staff_lc), None)
        share_pct = share_row.pct if share_row else 100.0
        val = calc_prepaid_card_bonus(r.item_name or "", bonuses)
        if val > 0:
            prepaid_card_bonus_sum += val * (share_pct / 100)
    prepaid_card_bonus_sum = round(prepaid_card_bonus_sum, 2)
    
    return {
        "staff_name": staff_name,
        "month": month,
        "base_salary": base_salary,
        "total_service": round(total_service, 2),
        "total_retail": round(total_retail, 2),
        "ratio": monthly_bonus["ratio"],
        "pct": monthly_bonus["pct"],
        "efficiency_bonus": monthly_bonus["amount"],
        "retail_commission": retail_comm,
        "prepaid_card_bonus": prepaid_card_bonus_sum,
        "total_earned": monthly_bonus["amount"] + retail_comm + prepaid_card_bonus_sum,
        "details": details,
        "config_multipliers": cfg["staff_monthly_multipliers"]
    }

@api.get("/incentives/monthly")
async def incentives_monthly(month: str):
    await _auto_sync_staff_from_pos()
    cfg = await get_config()
    async with async_session() as session:
        staff_list = (await session.execute(
            select(Staff).where(func.lower(Staff.role) != "owner", ~func.lower(Staff.name).contains("abhimanyu"))
        )).scalars().all()
    out = []
    for s in staff_list:
        rev = await _staff_month_revenue(s.name, month)
        service = rev["service"]
        retail = rev["retail"]

        # Calculate prepaid card sale bonuses
        async with async_session() as session2:
            tx_result = await session2.execute(
                select(POSTransaction.item_name, POSTransactionStaff.pct)
                .join(POSTransactionStaff)
                .where(
                    POSTransaction.date.like(f"{month}%"),
                    func.lower(func.trim(POSTransactionStaff.name)) == s.name.strip().lower()
                )
            )
            prepaid_card_bonus_sum = 0.0
            bonuses = cfg.get("prepaid_card_bonuses", [])
            for row in tx_result.all():
                item_name = row[0] or ""
                share_pct = row[1] or 100.0
                val = calc_prepaid_card_bonus(item_name, bonuses)
                if val > 0:
                    prepaid_card_bonus_sum += val * (share_pct / 100)
            prepaid_card_bonus_sum = round(prepaid_card_bonus_sum, 2)

        monthly = calc_monthly_bonus(service, s.base_salary, cfg["staff_monthly_multipliers"])
        retail_comm = round(retail * (cfg.get("retail_commission_pct", 0) / 100), 2)
        out.append({
            "staff_id": s.id, "staff_name": s.name, "base_salary": s.base_salary,
            "monthly_service_revenue": round(service, 2), "monthly_retail_revenue": round(retail, 2),
            "ratio": monthly["ratio"], "pct": monthly["pct"],
            "efficiency_bonus": monthly["amount"], "retail_commission": retail_comm,
            "prepaid_card_bonus": prepaid_card_bonus_sum,
            "total": monthly["amount"] + retail_comm + prepaid_card_bonus_sum,
        })
    return {"month": month, "config": cfg["staff_monthly_multipliers"], "rows": out}

@api.get("/incentives/manager")
async def incentives_manager(month: str):
    cfg = await get_config()
    async with async_session() as session:
        result = await session.execute(
            select(func.sum(POSTransaction.net_price)).where(POSTransaction.date.like(f"{month}%"))
        )
        revenue = result.scalar() or 0.0
    bonus = calc_manager_bonus(revenue, cfg["manager_milestones"])
    return {"month": month, "revenue": bonus["month_revenue"], "milestone": bonus["milestone"],
            "bonus_per_manager": bonus["bonus"], "milestones": cfg["manager_milestones"]}

# ------------------ Product Incentive Custom Mappings & Flagged Products ------------------

# ------------------ Vendors ------------------
@api.get("/vendors")
async def list_vendors():
    async with async_session() as session:
        result = await session.execute(select(Vendor).order_by(Vendor.name))
        return [v.to_dict() for v in result.scalars().all()]

@api.post("/vendors")
async def upsert_vendor(payload: VendorIn, user: dict = Depends(require_role("owner", "admin"))):
    async with async_session() as session:
        result = await session.execute(select(Vendor).where(Vendor.name == payload.name.strip()))
        existing = result.scalar_one_or_none()
        if existing:
            existing.lead_time_days = payload.lead_time_days
            existing.contact = payload.contact or ""
            existing.email = payload.email or ""
            existing.phone = payload.phone or ""
            existing.address = payload.address or ""
            existing.gst_number = payload.gst_number or ""
            existing.notes = payload.notes or ""
            await session.commit()
            return {"ok": True, "id": existing.id, "updated": True}
        doc = Vendor(
            id=new_id(), name=payload.name.strip(), lead_time_days=payload.lead_time_days,
            contact=payload.contact or "",
            email=payload.email or "",
            phone=payload.phone or "",
            address=payload.address or "",
            gst_number=payload.gst_number or "",
            notes=payload.notes or "", created_at=now_utc(),
        )
        session.add(doc)
        await session.commit()
        return {"ok": True, "id": doc.id, "updated": False}

@api.delete("/vendors/{vendor_id}")
async def delete_vendor(vendor_id: str, user: dict = Depends(require_role("owner", "admin"))):
    async with async_session() as session:
        result = await session.execute(select(Vendor).where(Vendor.id == vendor_id))
        vendor = result.scalar_one_or_none()
        if not vendor:
            raise HTTPException(404, "Vendor not found")
        await session.delete(vendor)
        await session.commit()
        return {"ok": True, "deleted": vendor_id}

# ------------------ Excel Template & Importers ------------------
def _make_header(ws, headers, note=None):
    if note:
        ws.cell(row=1, column=1, value=note).font = Font(italic=True, color="808080")
        start = 3
    else:
        start = 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=start, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F1F1F")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = max(14, len(str(h)) + 4)
    return start

@api.get("/inventory/template.xlsx")
async def inventory_template():
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "Instructions"
    tmpl_lines = [
        ("LUXURY SALON SUITE — Opening Stock & Master Data Template", True),
        ("", False),
        ("Fill each tab and upload via UI:", False),
        ("  1) 'Vendors' → POST /api/vendors/upload  (or fill inline)", False),
        ("  2) 'Opening Stock' → POST /api/inventory/opening-stock/upload", False),
        ("  3) 'Purchase Invoice' (repeatable) → POST /api/inventory/purchase-invoice/upload", False),
        ("", False),
        ("Rules:", False),
        ("• Locations are 3: STOREROOM, SERVICE FLOOR, RETAIL DISPLAY.", False),
        ("• POS product sales auto-deduct from RETAIL DISPLAY (FIFO by expiry).", False),
        ("• Service usage is checked out manually from STORE → SERVICE FLOOR.", False),
        ("• Every batch has its own expiry — near-expiry stock will be flagged.", False),
    ]
    for i, (txt, bold) in enumerate(tmpl_lines, 1):
        c = ws0.cell(row=i, column=1, value=txt)
        if bold:
            c.font = Font(bold=True, size=14, color="D4AF37")
    ws0.column_dimensions["A"].width = 110
    ws1 = wb.create_sheet("Vendors")
    _make_header(ws1, ["Vendor Name", "Lead Time (days)", "Contact", "Notes"],
                 note="One row per vendor. Lead time affects reorder point.")
    ws2 = wb.create_sheet("Opening Stock")
    _make_header(ws2, ["Vendor", "Category", "Product Name", "MOQ", "Unit Cost (₹)", "MRP (₹)",
                        "Storeroom Qty", "Service Floor Qty", "Retail Display Qty",
                        "Expiry Date (YYYY-MM-DD)", "Batch Number"],
                 note="One row per SKU batch.")
    ws3 = wb.create_sheet("Purchase Invoice")
    _make_header(ws3, ["Invoice #", "Vendor", "Invoice Date (YYYY-MM-DD)", "Product Name",
                        "Quantity", "Unit Cost (₹)", "Expiry Date (YYYY-MM-DD)", "Batch Number", "Notes"],
                 note="Repeat Invoice # + Vendor + Date on each line.")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=LSS_Inventory_Template.xlsx"},
    )

def _s(v):
    if v is None: return ""
    if isinstance(v, (datetime, date)): return v.strftime("%Y-%m-%d")
    return str(v).strip()

def _f(v):
    try:
        if v is None or v == "": return 0.0
        return float(v)
    except (ValueError, TypeError): return 0.0

def _read_sheet_dicts(wb, sheet_name):
    if sheet_name not in wb.sheetnames: return []
    ws = wb[sheet_name]
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first = _s(row[0]).lower()
        if first in {"vendor", "vendor name", "invoice #", "product name"}:
            header_row = i
            headers = [_s(c) for c in row]
            break
    if not header_row: return []
    out = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(c is None or _s(c) == "" for c in row): continue
        rec = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        out.append(rec)
    return out

def _parse_ledger_workbook(content: bytes, ledger: str):
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    header_idx = 0
    for i, row in enumerate(rows):
        if row and any(str(c or "").strip().lower() in ("product name", "description", "brand", "vendor") for c in row):
            header_idx = i
            break
    headers = [str(h or "").strip() for h in rows[header_idx]]
    def col(row, *names):
        for n in names:
            for i, h in enumerate(headers):
                if h.lower() == n.lower():
                    return row[i] if i < len(row) else None
        return None
    parsed = []
    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row): continue
        pn = str(col(row, "PRODUCT NAME", "Product Name") or "").strip()
        desc = str(col(row, "DESCRIPTION", "Description") or "").strip()
        name = (pn + " " + desc).strip() if pn else desc
        if not name: continue
        qty = _f(col(row, "UNITS", "Units", "Quantity", "Qty", "Storeroom Qty", "Retail Display Qty"))
        if qty <= 0: continue
        unit_price = _f(col(row, "UNIT mrp", "UNIT MRP", "Unit MRP", "MRP", "MRP (₹)"))
        unit_cost = _f(col(row, "UNIT COST", "Unit Cost", "Unit Cost (₹)", "COST"))
        vendor = str(col(row, "Vendor", "VENDOR") or "").strip()
        brand = str(col(row, "Brand", "BRAND") or "").strip()
        category = str(col(row, "Category", "CATEGORY") or "").strip() or "Uncategorized"
        moq = str(col(row, "MOQ", "Moq") or "").strip()
        exp = col(row, "EXPIRY", "Expiry", "Expiry Date")
        if isinstance(exp, (datetime, date)): exp = exp.strftime("%Y-%m-%d")
        elif exp: exp = str(exp)[:10]
        else: exp = None
        if ledger == "technical": exp = None
        parsed.append({
            "name": name, "vendor": vendor or brand, "brand": brand,
            "category": category, "qty": qty, "unit_price": unit_price,
            "unit_cost": unit_cost, "expiry": exp, "moq": moq,
        })
    return parsed

async def _import_ledger(content: bytes, ledger: str):
    loc = "retail" if ledger == "retail" else "floor"
    qty_field = "retail_qty" if ledger == "retail" else "floor_qty"
    parsed = _parse_ledger_workbook(content, ledger)
    created = 0
    batches_added = 0
    async with async_session() as session:
        for p in parsed:
            batch = SKUBatch(
                id=new_id(), qty=p["qty"], location=loc,
                expiry_date=p["expiry"], unit_cost=p["unit_cost"],
                invoice_id=None, received_at=now_utc(),
            )
            result = await session.execute(select(SKU).where(SKU.name == p["name"], SKU.ledger == ledger))
            existing = result.scalar_one_or_none()
            if existing:
                batch.sku_id = existing.id
                session.add(batch)
                setattr(existing, qty_field, (getattr(existing, qty_field) or 0) + p["qty"])
                batches_added += 1
            else:
                sku_id = new_id()
                sku = SKU(
                    id=sku_id, name=p["name"], category=p["category"],
                    ledger=ledger, vendor_name=p["vendor"], brand=p["brand"],
                    moq=p["moq"], unit_cost=p["unit_cost"], unit_price=p["unit_price"],
                    store_qty=0,
                    floor_qty=p["qty"] if ledger == "technical" else 0,
                    retail_qty=p["qty"] if ledger == "retail" else 0,
                    created_at=now_utc(),
                )
                batch.sku_id = sku_id
                session.add(sku)
                session.add(batch)
                created += 1
        await session.commit()
    return {"ledger": ledger, "skus_created": created, "batches_added": batches_added, "rows": len(parsed)}

@api.post("/inventory/import/open-stock")
async def import_open_stock(file: UploadFile = File(...), ledger: str = Query("retail"),
                            user: dict = Depends(require_role("owner", "admin"))):
    if ledger not in ("retail", "technical"):
        raise HTTPException(400, "ledger must be 'retail' or 'technical'")
    data = await file.read()
    try:
        res = await _import_ledger(data, ledger)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse Excel: {e}")
    if res["rows"] == 0:
        raise HTTPException(400, "No valid rows found.")
    return res

@api.post("/vendors/upload")
async def vendors_upload(file: UploadFile = File(...), user: dict = Depends(require_role("owner", "admin"))):
    data = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    rows = _read_sheet_dicts(wb, "Vendors")
    if not rows:
        raise HTTPException(400, "No rows in 'Vendors' sheet")
    imported = 0
    async with async_session() as session:
        for r in rows:
            name = _s(r.get("Vendor Name") or r.get("Vendor"))
            if not name: continue
            result = await session.execute(select(Vendor).where(Vendor.name == name))
            existing = result.scalar_one_or_none()
            if existing:
                existing.lead_time_days = int(_f(r.get("Lead Time (days)"))) or 4
                existing.contact = _s(r.get("Contact"))
                existing.notes = _s(r.get("Notes"))
            else:
                session.add(Vendor(
                    id=new_id(), name=name,
                    lead_time_days=int(_f(r.get("Lead Time (days)"))) or 4,
                    contact=_s(r.get("Contact")), notes=_s(r.get("Notes")),
                    created_at=now_utc(),
                ))
            imported += 1
        await session.commit()
    return {"imported": imported}

@api.post("/inventory/opening-stock/upload")
async def opening_stock_upload(file: UploadFile = File(...), user: dict = Depends(require_role("owner", "admin"))):
    data = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    rows = _read_sheet_dicts(wb, "Opening Stock")
    if not rows:
        raise HTTPException(400, "No rows in 'Opening Stock' sheet")
    imported = 0
    async with async_session() as session:
        for r in rows:
            name = _s(r.get("Product Name"))
            if not name: continue
            vendor_name = _s(r.get("Vendor"))
            store_q = _f(r.get("Storeroom Qty"))
            floor_q = _f(r.get("Service Floor Qty"))
            retail_q = _f(r.get("Retail Display Qty"))
            unit_cost = _f(r.get("Unit Cost (₹)"))
            expiry = _s(r.get("Expiry Date (YYYY-MM-DD)")) or None
            if expiry and len(expiry) > 10: expiry = expiry[:10]
            batch_no = _s(r.get("Batch Number"))

            vendor_id = None
            if vendor_name:
                v_result = await session.execute(select(Vendor).where(Vendor.name == vendor_name))
                v = v_result.scalar_one_or_none()
                if v: vendor_id = v.id

            new_batches = []
            for qty_val, loc in [(store_q, "store"), (floor_q, "floor"), (retail_q, "retail")]:
                if qty_val > 0:
                    new_batches.append(SKUBatch(
                        id=new_id(), qty=qty_val, location=loc, expiry_date=expiry,
                        unit_cost=unit_cost, invoice_id=None, batch_number=batch_no, received_at=now_utc(),
                    ))

            result = await session.execute(select(SKU).where(SKU.name == name))
            existing = result.scalar_one_or_none()
            if existing:
                for b in new_batches:
                    b.sku_id = existing.id
                    session.add(b)
                existing.store_qty = (existing.store_qty or 0) + store_q
                existing.floor_qty = (existing.floor_qty or 0) + floor_q
                existing.retail_qty = (existing.retail_qty or 0) + retail_q
                if vendor_id: existing.vendor_id = vendor_id
                if vendor_name: existing.vendor_name = vendor_name
                if unit_cost: existing.unit_cost = unit_cost
                mrp = _f(r.get("MRP (₹)"))
                if mrp: existing.unit_price = mrp
                cat = _s(r.get("Category"))
                if cat: existing.category = cat
            else:
                sku_id = new_id()
                sku = SKU(
                    id=sku_id, name=name,
                    category=_s(r.get("Category")) or "Uncategorized",
                    ledger=_s(r.get("Ledger")).lower() if r.get("Ledger") else "retail",
                    vendor_id=vendor_id, vendor_name=vendor_name or "",
                    moq=_s(r.get("MOQ")), unit_cost=unit_cost,
                    unit_price=_f(r.get("MRP (₹)")),
                    store_qty=store_q, floor_qty=floor_q, retail_qty=retail_q,
                    created_at=now_utc(),
                )
                session.add(sku)
                for b in new_batches:
                    b.sku_id = sku_id
                    session.add(b)
            imported += 1
        await session.commit()
    return {"imported": imported}

@api.post("/inventory/purchase-invoice/upload")
async def purchase_invoice_upload(file: UploadFile = File(...), user: dict = Depends(require_role("owner", "admin", "manager"))):
    data = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    rows = _read_sheet_dicts(wb, "Purchase Invoice")
    if not rows:
        raise HTTPException(400, "No rows in 'Purchase Invoice' sheet")
    grouped: Dict[str, List[dict]] = defaultdict(list)
    meta: Dict[str, dict] = {}
    for r in rows:
        inv = _s(r.get("Invoice #"))
        if not inv: continue
        grouped[inv].append(r)
        if inv not in meta:
            meta[inv] = {
                "vendor": _s(r.get("Vendor")),
                "invoice_date": _s(r.get("Invoice Date (YYYY-MM-DD)"))[:10] or datetime.now().strftime("%Y-%m-%d"),
            }
    created = 0
    async with async_session() as session:
        for inv_num, lines_data in grouped.items():
            pi_id = new_id()
            total = 0.0
            processed = []
            for line in lines_data:
                name = _s(line.get("Product Name"))
                result = await session.execute(select(SKU).where(SKU.name == name))
                sku = result.scalar_one_or_none()
                if not sku: continue
                qty = _f(line.get("Quantity"))
                uc = _f(line.get("Unit Cost (₹)"))
                expiry = _s(line.get("Expiry Date (YYYY-MM-DD)"))[:10] or None
                session.add(SKUBatch(
                    id=new_id(), sku_id=sku.id, qty=qty, location="store",
                    expiry_date=expiry, unit_cost=uc, invoice_id=pi_id,
                    batch_number=_s(line.get("Batch Number")), received_at=now_utc(),
                ))
                sku.store_qty = (sku.store_qty or 0) + qty
                sku.unit_cost = uc
                total += qty * uc
                processed.append({"sku_id": sku.id, "sku_name": name, "quantity": qty,
                                  "unit_cost": uc, "expiry_date": expiry, "line_total": qty * uc})
            if processed:
                pi = PurchaseInvoice(
                    id=pi_id, invoice_number=inv_num,
                    vendor=meta[inv_num]["vendor"], invoice_date=meta[inv_num]["invoice_date"],
                    total=round(total, 2), notes="",
                    created_by=user["email"], created_at=now_utc(),
                )
                session.add(pi)
                for pl in processed:
                    session.add(PurchaseInvoiceLine(invoice_id=pi_id, **pl))
                created += 1
        await session.commit()
    return {"invoices_created": created}

# ------------------ Reports ------------------
async def _cumulative_data(date_from: str, date_to: str, only_unpaid: bool):
    cfg = await get_config()
    async with async_session() as session:
        staff_list = (await session.execute(select(Staff))).scalars().all()
        dates_result = await session.execute(
            select(distinct(POSTransaction.date)).where(
                POSTransaction.date >= date_from, POSTransaction.date <= date_to
            ).order_by(POSTransaction.date)
        )
        dates = [r[0] for r in dates_result.all()]
    rows = []
    for s in staff_list:
        total_service = total_bonus = total_product = 0.0
        unpaid_days = []
        paid_days = []
        for day in dates:
            rev = await _staff_day_revenue(s.name, day)
            b = calc_daily_bonus(rev["service"], cfg["staff_daily_tiers"])
            pi = await _staff_day_product_incentive(s.name, day, cfg.get("product_incentives", []))
            day_total = b["bonus"] + pi
            async with async_session() as session2:
                payout = (await session2.execute(
                    select(Payout).where(Payout.staff_id == s.id, Payout.payout_date == day)
                )).scalar_one_or_none()
            if day_total <= 0: continue
            if payout:
                paid_days.append({"day": day, "amount": day_total})
            else:
                unpaid_days.append({"day": day, "service": rev["service"], "bonus": b["bonus"],
                                   "product_incentive": pi, "total": day_total})
                total_service += rev["service"]
                total_bonus += b["bonus"]
                total_product += pi
        if not only_unpaid or unpaid_days:
            rows.append({
                "staff_id": s.id, "staff_name": s.name,
                "unpaid_days_count": len(unpaid_days), "unpaid_days": unpaid_days,
                "cumulative_service": round(total_service, 2),
                "cumulative_bonus": round(total_bonus, 2),
                "cumulative_product_incentive": round(total_product, 2),
                "cumulative_total_due": round(total_bonus + total_product, 2),
                "paid_days_count": len(paid_days),
            })
    return {"from": date_from, "to": date_to, "rows": rows}

@api.get("/reports/cumulative-payouts")
async def cumulative_payouts(date_from: str, date_to: str, only_unpaid: bool = True):
    return await _cumulative_data(date_from, date_to, only_unpaid)

@api.get("/reports/monthly-incentives.xlsx")
async def export_monthly_incentives_xlsx(month: str):
    data = await incentives_monthly(month)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Monthly Report {month}"
    _make_header(ws, [
        "Staff Name", "Base Salary ₹", "Monthly Service Revenue ₹",
        "Efficiency Ratio (×)", "Slab %", "Efficiency Bonus ₹",
        "Retail Commission ₹", "Prepaid Card Bonus ₹", "Total Earned ₹"
    ], note=f"Geetanjali Salon Monthly Staff Incentive & Efficiency Report — {month}")
    
    rows = data.get("rows", [])
    for i, r in enumerate(rows, start=4):
        ws.cell(i, 1, r.get("staff_name", ""))
        ws.cell(i, 2, r.get("base_salary", 0))
        ws.cell(i, 3, r.get("monthly_service_revenue", 0))
        ws.cell(i, 4, f"{r.get('ratio', 0)}×")
        ws.cell(i, 5, f"{r.get('pct', 0)}%")
        ws.cell(i, 6, r.get("efficiency_bonus", 0))
        ws.cell(i, 7, r.get("retail_commission", 0))
        ws.cell(i, 8, r.get("prepaid_card_bonus", 0))
        ws.cell(i, 9, r.get("total", 0))

    if rows:
        tot_row = len(rows) + 4
        ws.cell(tot_row, 1, "TOTAL")
        ws.cell(tot_row, 2, sum(r.get("base_salary", 0) for r in rows))
        ws.cell(tot_row, 3, sum(r.get("monthly_service_revenue", 0) for r in rows))
        ws.cell(tot_row, 4, "—")
        ws.cell(tot_row, 5, "—")
        ws.cell(tot_row, 6, sum(r.get("efficiency_bonus", 0) for r in rows))
        ws.cell(tot_row, 7, sum(r.get("retail_commission", 0) for r in rows))
        ws.cell(tot_row, 8, sum(r.get("prepaid_card_bonus", 0) for r in rows))
        ws.cell(tot_row, 9, sum(r.get("total", 0) for r in rows))

        font_bold = Font(name="Calibri", size=11, bold=True)
        for col_idx in range(1, 10):
            cell = ws.cell(tot_row, col_idx)
            cell.font = font_bold

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Geetanjali_Monthly_Incentives_{month}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

@api.get("/reports/cumulative-payouts.xlsx")
async def cumulative_payouts_xlsx(date_from: str, date_to: str, only_unpaid: bool = True):
    data = await _cumulative_data(date_from, date_to, only_unpaid)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    _make_header(ws, ["Staff", "Unpaid Days", "Cumulative Service ₹", "Cumulative Bonus ₹",
                       "Cumulative Product Incentive ₹", "Total Due ₹"],
                 note=f"Cumulative unpaid incentives {data['from']} → {data['to']}")
    for i, r in enumerate(data["rows"], start=4):
        ws.cell(i, 1, r["staff_name"]); ws.cell(i, 2, r["unpaid_days_count"])
        ws.cell(i, 3, r["cumulative_service"]); ws.cell(i, 4, r["cumulative_bonus"])
        ws.cell(i, 5, r["cumulative_product_incentive"]); ws.cell(i, 6, r["cumulative_total_due"])
    ws2 = wb.create_sheet("Day-wise Unpaid")
    _make_header(ws2, ["Staff", "Date", "Service ₹", "Daily Bonus ₹", "Product Incentive ₹", "Total ₹"])
    i = 4
    for r in data["rows"]:
        for d in r["unpaid_days"]:
            ws2.cell(i, 1, r["staff_name"]); ws2.cell(i, 2, d["day"])
            ws2.cell(i, 3, d["service"]); ws2.cell(i, 4, d["bonus"])
            ws2.cell(i, 5, d["product_incentive"]); ws2.cell(i, 6, d["total"])
            i += 1
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"LSS_Cumulative_{date_from}_to_{date_to}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})

@api.get("/config/product-incentives.xlsx")
async def export_product_incentives():
    cfg = await get_config()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Incentives"
    _make_header(ws, ["Brand", "Product Pattern", "Min Price ₹", "Max Price ₹", "Incentive ₹"],
                 note="Product incentive rules.")
    for i, r in enumerate(cfg.get("product_incentives", []), start=4):
        ws.cell(i, 1, r.get("brand", "")); ws.cell(i, 2, r.get("pattern", ""))
        ws.cell(i, 3, r.get("min_price", "")); ws.cell(i, 4, r.get("max_price", ""))
        ws.cell(i, 5, r.get("amount", 0))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=LSS_Product_Incentives.xlsx"})

# ------------------ Payouts ------------------
@api.post("/payouts/confirm")
async def confirm_payout(payload: ConfirmPayoutIn, user: dict = Depends(require_role("owner", "admin"))):
    async with async_session() as session:
        result = await session.execute(
            select(Payout).where(Payout.staff_id == payload.staff_id, Payout.payout_date == payload.payout_date)
        )
        existing = result.scalar_one_or_none()
        payout_id = new_id()
        if existing:
            existing.amount = payload.amount
            existing.breakdown = payload.breakdown
            existing.confirmed_by = user["email"]
            existing.confirmed_at = now_utc()
            payout_id = existing.id
        else:
            p = Payout(
                id=payout_id, staff_id=payload.staff_id, payout_date=payload.payout_date,
                amount=payload.amount, breakdown=payload.breakdown,
                confirmed_by=user["email"], confirmed_at=now_utc(),
            )
            session.add(p)
        await session.commit()
    return {"ok": True, "payout": {"id": payout_id, "staff_id": payload.staff_id,
            "payout_date": payload.payout_date, "amount": payload.amount,
            "breakdown": payload.breakdown, "confirmed_by": user["email"], "confirmed_at": now_utc()}}

@api.get("/payouts")
async def list_payouts(day: Optional[str] = None):
    async with async_session() as session:
        q = select(Payout)
        if day:
            q = q.where(Payout.payout_date == day)
        q = q.order_by(Payout.confirmed_at.desc()).limit(500)
        result = await session.execute(q)
        return [p.to_dict() for p in result.scalars().all()]

def calc_prepaid_card_bonus(item_name: str, bonuses: list) -> float:
    if not item_name or not bonuses:
        return 0.0
    item_name_lower = item_name.lower()
    for b in bonuses:
        pat = (b.get("pattern") or "").strip().lower()
        if pat and pat in item_name_lower:
            return float(b.get("amount") or 0.0)
    return 0.0

@api.get("/payouts/monthly")
async def monthly_payouts(month: str):
    await _auto_sync_staff_from_pos()
    cfg = await get_config()
    async with async_session() as session:
        staff_list = (await session.execute(
            select(Staff).where(func.lower(Staff.role) != "owner", ~func.lower(Staff.name).contains("abhimanyu")).order_by(Staff.name)
        )).scalars().all()
    
    out = []
    for s in staff_list:
        staff_lc = s.name.strip().lower()
        async with async_session() as session2:
            q = (
                select(POSTransaction.type, func.sum(POSTransaction.net_price * POSTransactionStaff.pct / 100))
                .join(POSTransactionStaff)
                .where(
                    POSTransaction.date.like(f"{month}%"),
                    func.lower(func.trim(POSTransactionStaff.name)) == staff_lc
                )
                .group_by(POSTransaction.type)
            )
            result = await session2.execute(q)
            service_rev = 0.0
            retail_rev = 0.0
            for row in result.all():
                t_name = str(row[0]).strip().lower()
                if t_name == "service":
                    service_rev = row[1] or 0.0
                elif t_name in ("product", "retail"):
                    retail_rev = row[1] or 0.0

            dates_result = await session2.execute(
                select(distinct(POSTransaction.date)).where(POSTransaction.date.like(f"{month}%"))
            )
            dates = [r[0] for r in dates_result.all()]
            
            daily_bonus_sum = 0.0
            product_inc_sum = 0.0
            for d in dates:
                d_rev = await _staff_day_revenue(s.name, d)
                b = calc_daily_bonus(d_rev["service"], cfg["staff_daily_tiers"])
                pi = await _staff_day_product_incentive(s.name, d, cfg.get("product_incentives", []))
                daily_bonus_sum += b["bonus"]
                product_inc_sum += pi

            # Calculate prepaid card sale bonuses
            tx_result = await session2.execute(
                select(POSTransaction.item_name, POSTransactionStaff.pct)
                .join(POSTransactionStaff)
                .where(
                    POSTransaction.date.like(f"{month}%"),
                    func.lower(func.trim(POSTransactionStaff.name)) == staff_lc
                )
            )
            prepaid_card_bonus_sum = 0.0
            bonuses = cfg.get("prepaid_card_bonuses", [])
            for row in tx_result.all():
                item_name = row[0] or ""
                share_pct = row[1] or 100.0
                val = calc_prepaid_card_bonus(item_name, bonuses)
                if val > 0:
                    prepaid_card_bonus_sum += val * (share_pct / 100)
            prepaid_card_bonus_sum = round(prepaid_card_bonus_sum, 2)

            monthly_tier = calc_monthly_bonus(service_rev, s.base_salary, cfg["staff_monthly_multipliers"])
            legacy_pct = cfg.get("retail_commission_pct", 0)
            retail_comm = round(retail_rev * (legacy_pct / 100), 2) if legacy_pct else 0.0

            total_earned = (s.base_salary or 25000) + monthly_tier["amount"] + daily_bonus_sum + product_inc_sum + retail_comm + prepaid_card_bonus_sum

            payout_result = await session2.execute(
                select(Payout).where(Payout.staff_id == s.id, Payout.payout_date == month)
            )
            payout = payout_result.scalar_one_or_none()

            out.append({
                "staff_id": s.id,
                "staff_name": s.name,
                "base_salary": s.base_salary or 25000,
                "service_revenue": round(service_rev, 2),
                "retail_revenue": round(retail_rev, 2),
                "monthly_bonus": monthly_tier["amount"],
                "daily_bonus_sum": round(daily_bonus_sum, 2),
                "product_incentive": round(product_inc_sum, 2),
                "retail_commission": retail_comm,
                "prepaid_card_bonus": prepaid_card_bonus_sum,
                "total_earned": round(total_earned, 2),
                "confirmed": bool(payout),
                "confirmed_at": payout.confirmed_at if payout else None,
                "confirmed_by": payout.confirmed_by if payout else None,
            })
    return {"month": month, "rows": out}


# ------------------ Inventory ------------------
@api.get("/inventory/skus")
async def list_skus(ledger: Optional[str] = None):
    async with async_session() as session:
        q = select(SKU).options(selectinload(SKU.batches))
        if ledger in ("retail", "technical"):
            q = q.where(SKU.ledger == ledger)
        q = q.order_by(SKU.name)
        result = await session.execute(q)
        docs = []
        for s in result.scalars().unique().all():
            d = s.to_dict()
            d.setdefault("retail_qty", 0)
            d.setdefault("store_qty", 0)
            d.setdefault("floor_qty", 0)
            d.setdefault("ledger", "retail")
            docs.append(d)
        return docs

@api.post("/inventory/checkout")
async def inventory_checkout(payload: CheckoutIn, user: dict = Depends(get_current_user)):
    async with async_session() as session:
        result = await session.execute(select(SKU).where(SKU.id == payload.sku_id))
        sku = result.scalar_one_or_none()
        if not sku:
            raise HTTPException(404, "SKU not found")
        if (sku.store_qty or 0) < payload.quantity:
            raise HTTPException(400, "Insufficient store stock")
        moved = await _move_batches_session(session, sku.id, "store", "floor", payload.quantity)
        if moved <= 0:
            raise HTTPException(400, "Insufficient store batches")
        session.add(Checkout(
            id=new_id(), sku_id=payload.sku_id, sku_name=sku.name,
            quantity=moved, notes=payload.notes or "", source="manual",
            checked_out_by=user["email"], checked_out_at=now_utc(),
        ))
        await session.commit()
    return {"ok": True, "moved": moved}

@api.post("/inventory/consume")
async def inventory_consume(payload: CheckoutIn, user: dict = Depends(get_current_user)):
    async with async_session() as session:
        result = await session.execute(select(SKU).where(SKU.id == payload.sku_id))
        sku = result.scalar_one_or_none()
        if not sku:
            raise HTTPException(404, "SKU not found")
        if (sku.floor_qty or 0) < payload.quantity:
            raise HTTPException(400, "Insufficient floor stock")
        consumed = await _consume_batches_session(session, sku.id, "floor", payload.quantity)
        if consumed <= 0:
            raise HTTPException(400, "Insufficient floor batches")
        session.add(Checkout(
            id=new_id(), sku_id=payload.sku_id, sku_name=sku.name,
            quantity=consumed, notes=payload.notes or "", source="service-use",
            checked_out_by=user["email"], checked_out_at=now_utc(),
        ))
        await session.commit()
    return {"ok": True, "consumed": consumed}

@api.post("/inventory/receive")
async def inventory_receive(payload: ReceiveIn, user: dict = Depends(require_role("owner", "admin"))):
    async with async_session() as session:
        result = await session.execute(select(SKU).where(SKU.id == payload.sku_id))
        sku = result.scalar_one_or_none()
        if not sku:
            raise HTTPException(404, "SKU not found")
        ledger = sku.ledger or "retail"
        loc = "retail" if ledger == "retail" else "store"
        qty_field = "retail_qty" if ledger == "retail" else "store_qty"
        session.add(SKUBatch(
            id=new_id(), sku_id=payload.sku_id, qty=payload.quantity, location=loc,
            expiry_date=payload.expiry_date or None,
            unit_cost=payload.unit_cost or sku.unit_cost or 0,
            invoice_id=None, received_at=now_utc(),
        ))
        setattr(sku, qty_field, (getattr(sku, qty_field) or 0) + payload.quantity)
        await session.commit()
    return {"ok": True}

@api.post("/inventory/skus")
async def create_sku(payload: SKUCreateIn, user: dict = Depends(require_role("owner", "admin"))):
    async with async_session() as session:
        result = await session.execute(select(SKU).where(SKU.name == payload.name.strip()))
        if result.scalar_one_or_none():
            raise HTTPException(400, "SKU with this name already exists")
        sku_id = new_id()
        sku = SKU(
            id=sku_id, name=payload.name.strip(), category=payload.category,
            unit_cost=payload.unit_cost, unit_price=payload.unit_price,
            store_qty=payload.opening_store_qty, floor_qty=payload.opening_floor_qty,
            created_at=now_utc(),
        )
        session.add(sku)
        if payload.opening_store_qty > 0:
            session.add(SKUBatch(
                id=new_id(), sku_id=sku_id, qty=payload.opening_store_qty, location="store",
                expiry_date=payload.opening_expiry, unit_cost=payload.unit_cost, received_at=now_utc(),
            ))
        if payload.opening_floor_qty > 0:
            session.add(SKUBatch(
                id=new_id(), sku_id=sku_id, qty=payload.opening_floor_qty, location="floor",
                expiry_date=payload.opening_expiry, unit_cost=payload.unit_cost, received_at=now_utc(),
            ))
        await session.commit()
    return {"ok": True, "id": sku_id}

@api.post("/inventory/purchase-invoice")
async def create_purchase_invoice(payload: PurchaseInvoiceIn, user: dict = Depends(require_role("owner", "admin", "manager"))):
    pi_id = new_id()
    total = 0.0
    processed_lines = []
    async with async_session() as session:
        for line in payload.lines:
            result = await session.execute(select(SKU).where(SKU.id == line.sku_id))
            sku = result.scalar_one_or_none()
            if not sku:
                raise HTTPException(400, f"SKU {line.sku_id} not found")
            session.add(SKUBatch(
                id=new_id(), sku_id=line.sku_id, qty=line.quantity, location="store",
                expiry_date=line.expiry_date, unit_cost=line.unit_cost,
                invoice_id=pi_id, received_at=now_utc(),
            ))
            sku.store_qty = (sku.store_qty or 0) + line.quantity
            sku.unit_cost = line.unit_cost
            total += line.quantity * line.unit_cost
            processed_lines.append({
                "sku_id": line.sku_id, "sku_name": sku.name,
                "quantity": line.quantity, "unit_cost": line.unit_cost,
                "expiry_date": line.expiry_date, "line_total": line.quantity * line.unit_cost,
            })
        pi = PurchaseInvoice(
            id=pi_id, invoice_number=payload.invoice_number, vendor=payload.vendor,
            invoice_date=payload.invoice_date, total=round(total, 2),
            notes=payload.notes or "", created_by=user["email"], created_at=now_utc(),
        )
        session.add(pi)
        for pl in processed_lines:
            session.add(PurchaseInvoiceLine(invoice_id=pi_id, **pl))
        await session.commit()
    doc = pi.to_dict()
    return doc

@api.get("/inventory/purchase-invoices")
async def list_purchase_invoices(limit: int = 100):
    async with async_session() as session:
        q = select(PurchaseInvoice).options(selectinload(PurchaseInvoice.lines)).order_by(
            PurchaseInvoice.created_at.desc()
        ).limit(limit)
        result = await session.execute(q)
        return [pi.to_dict() for pi in result.scalars().unique().all()]

@api.get("/inventory/expiring")
async def expiring_batches(days: int = 60):
    cutoff = (datetime.now(timezone.utc) + timedelta(days=days)).strftime("%Y-%m-%d")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    async with async_session() as session:
        q = (
            select(SKUBatch, SKU.name, SKU.category)
            .join(SKU, SKUBatch.sku_id == SKU.id)
            .where(SKUBatch.expiry_date.isnot(None), SKUBatch.expiry_date <= cutoff, SKUBatch.qty > 0)
            .order_by(SKUBatch.expiry_date)
        )
        result = await session.execute(q)
        out = []
        for batch, sku_name, category in result.all():
            try:
                d1 = datetime.strptime(batch.expiry_date, "%Y-%m-%d")
                d0 = datetime.strptime(today, "%Y-%m-%d")
                days_left = (d1 - d0).days
            except Exception:
                days_left = 999
            out.append({
                "sku_id": batch.sku_id, "sku_name": sku_name, "category": category,
                "batch_id": batch.id, "location": batch.location, "quantity": batch.qty,
                "expiry_date": batch.expiry_date, "days_left": days_left,
                "unit_cost": batch.unit_cost or 0,
                "value_at_risk": round(batch.qty * (batch.unit_cost or 0), 2),
                "recommendation": "RETURN_TO_VENDOR" if days_left < 30 else "PROMOTE_OR_MONITOR",
            })
        return out

@api.get("/inventory/checkouts")
async def list_checkouts(limit: int = 100):
    async with async_session() as session:
        q = select(Checkout).order_by(Checkout.checked_out_at.desc()).limit(limit)
        result = await session.execute(q)
        return [c.to_dict() for c in result.scalars().all()]

@api.get("/inventory/purchase-orders")
async def purchase_orders():
    cfg = await get_config()
    default_lead = cfg["inventory"]["lead_time_days"]
    buf = cfg["inventory"]["safety_buffer_pct"] / 100
    async with async_session() as session:
        # Velocity: total checkouts per SKU / 30 days
        vel_q = select(Checkout.sku_id, func.sum(Checkout.quantity)).group_by(Checkout.sku_id)
        vel_result = await session.execute(vel_q)
        vel_map = {r[0]: r[1] / 30 for r in vel_result.all()}

        vendors_result = await session.execute(select(Vendor))
        vendors = {v.id: v for v in vendors_result.scalars().all()}

        skus_result = await session.execute(select(SKU))
        skus = skus_result.scalars().all()

    drafts = []
    for s in skus:
        v = vel_map.get(s.id, 0.3)
        vendor = vendors.get(s.vendor_id) if s.vendor_id else None
        lead = vendor.lead_time_days if vendor else default_lead
        reorder_point = round(v * lead * (1 + buf), 2)
        total_on_hand = (s.store_qty or 0) + (s.floor_qty or 0) + (s.retail_qty or 0)
        needs_reorder = total_on_hand < reorder_point
        drafts.append({
            "sku_id": s.id, "sku_name": s.name, "category": s.category,
            "vendor": s.vendor_name or (vendor.name if vendor else None),
            "lead_time_days": lead, "on_hand": total_on_hand,
            "velocity_per_day": round(v, 2), "reorder_point": reorder_point,
            "needs_reorder": needs_reorder,
            "suggested_order_qty": max(0, round(reorder_point * 2 - total_on_hand, 0)),
            "unit_cost": s.unit_cost,
            "estimated_cost": round(max(0, reorder_point * 2 - total_on_hand) * s.unit_cost, 2),
        })
    return drafts

# ------------------ Config ------------------
@api.get("/config")
async def get_config_ep():
    return await get_config()

@api.put("/config")
async def update_config(payload: ConfigUpdateIn, user: dict = Depends(require_role("owner", "admin"))):
    updates = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    async with async_session() as session:
        result = await session.execute(select(AppConfig).where(AppConfig.id == "master"))
        cfg = result.scalar_one_or_none()
        if cfg:
            data = dict(cfg.data or {})
            data.update(updates)
            cfg.data = data
        else:
            data = dict(DEFAULT_CONFIG)
            data.update(updates)
            session.add(AppConfig(id="master", data=data))
        await session.commit()
    return await get_config()

# ------------------ Owner Dashboard ------------------
@api.get("/dashboard/owner")
async def owner_dashboard():
    async with async_session() as session:
        # Working capital
        skus_result = await session.execute(select(SKU))
        skus = skus_result.scalars().all()
        working_capital = sum(
            ((s.store_qty or 0) + (s.floor_qty or 0) + (s.retail_qty or 0)) * (s.unit_cost or 0) for s in skus
        )
        # Revenue by type
        rev_result = await session.execute(
            select(POSTransaction.type, func.sum(POSTransaction.net_price), func.count())
            .group_by(POSTransaction.type)
        )
        service_rev = retail_rev = 0.0
        for row in rev_result.all():
            if str(row[0]).lower() == "service": service_rev = row[1] or 0
            elif str(row[0]).lower() == "product": retail_rev = row[1] or 0
        # Leakage
        co_result = await session.execute(select(func.sum(Checkout.quantity)))
        checkout_qty = co_result.scalar() or 0
        ps_result = await session.execute(
            select(func.sum(POSTransaction.quantity)).where(POSTransaction.type == "Product")
        )
        product_sold_qty = ps_result.scalar() or 0
        leakage = max(0, checkout_qty - product_sold_qty)
        # Quality alerts
        qf_result = await session.execute(
            select(func.count()).select_from(POSTransaction).where(POSTransaction.is_quality_failure == True)
        )
        qf_count = qf_result.scalar() or 0
        # Payouts pending
        dates_result = await session.execute(select(distinct(POSTransaction.date)))
        dates_count = len(dates_result.all())
        staff_count_result = await session.execute(select(func.count()).select_from(Staff))
        staff_count = staff_count_result.scalar() or 0
        paid_count_result = await session.execute(select(func.count()).select_from(Payout))
        paid_count = paid_count_result.scalar() or 0
        pending_payouts = max(0, dates_count * staff_count - paid_count)

    return {
        "working_capital": round(working_capital, 2),
        "total_service_revenue": round(service_rev, 2),
        "total_retail_revenue": round(retail_rev, 2),
        "inventory_leakage_units": leakage,
        "quality_alerts": qf_count,
        "pending_payouts": pending_payouts,
        "sku_count": len(skus),
        "staff_count": staff_count,
    }

# ══════════════════════════════════════════════════════════════
#  ENTERPRISE INVENTORY MODULE API ENDPOINTS
# ══════════════════════════════════════════════════════════════
class InventoryProductIn(BaseModel):
    product_code: Optional[str] = None
    barcode: Optional[str] = None
    name: str
    category: Optional[str] = "Retail"
    brand: Optional[str] = ""
    vendor_id: Optional[str] = None
    vendor_name: Optional[str] = ""
    unit: Optional[str] = "Piece"
    unit_cost: float = 0.0
    mrp: float = 0.0
    selling_price: float = 0.0
    min_stock: float = 5.0
    reorder_level: float = 10.0
    store_qty: float = 0.0
    floor_qty: float = 0.0
    retail_qty: float = 0.0

class PurchaseInPayload(BaseModel):
    method: str = "direct"  # "po" | "direct"
    vendor_id: Optional[str] = None
    vendor_name: Optional[str] = ""
    po_number: Optional[str] = None
    invoice_number: str
    invoice_date: str
    items: List[dict]  # [{ sku_id, quantity, purchase_price, mrp, gst_pct, batch_number, expiry_date }]
    remarks: Optional[str] = ""

class ManualStockOutPayload(BaseModel):
    sku_id: str
    quantity: float
    issued_to: str
    reason: str
    remarks: Optional[str] = ""

class POSServiceStockOutPayload(BaseModel):
    service_name: str
    pos_transaction_id: Optional[str] = None
    performed_by: Optional[str] = ""

@api.get("/inventory/master")
async def get_inventory_master(
    category: Optional[str] = None,
    brand: Optional[str] = None,
    vendor: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
):
    async with async_session() as session:
        q = select(SKU).options(selectinload(SKU.batches)).order_by(SKU.name)
        if category and category.lower() != "all":
            q = q.where(SKU.category == category)
        if brand:
            q = q.where(SKU.brand == brand)
        if vendor:
            q = q.where(SKU.vendor_name == vendor)

        result = await session.execute(q)
        skus = [s.to_dict() for s in result.scalars().unique().all()]

        if search:
            s_lower = search.lower()
            skus = [
                s for s in skus
                if s_lower in (s["name"] or "").lower()
                or s_lower in (s["product_code"] or "").lower()
                or s_lower in (s["barcode"] or "").lower()
            ]

        if status:
            if status.lower() == "low_stock":
                skus = [s for s in skus if 0 < s["current_stock"] <= s["min_stock"]]
            elif status.lower() == "out_of_stock":
                skus = [s for s in skus if s["current_stock"] == 0]
            elif status.lower() == "active":
                skus = [s for s in skus if s["current_stock"] > s["min_stock"]]

        return skus

@api.post("/inventory/product")
async def create_inventory_product(payload: InventoryProductIn, user: dict = Depends(require_role("owner", "admin", "manager"))):
    async with async_session() as session:
        # Check if product with same name exists
        existing_res = await session.execute(select(SKU).where(SKU.name == payload.name.strip()))
        existing = existing_res.scalar_one_or_none()
        if existing:
            existing.category = payload.category or existing.category
            existing.brand = payload.brand or existing.brand
            existing.vendor_name = payload.vendor_name or existing.vendor_name
            existing.unit = payload.unit or existing.unit
            existing.unit_cost = payload.unit_cost or existing.unit_cost
            existing.mrp = payload.mrp or existing.mrp
            existing.selling_price = payload.selling_price or existing.selling_price
            existing.store_qty = (existing.store_qty or 0) + (payload.store_qty or 0)
            existing.updated_at = now_utc()
            await session.commit()
            return existing.to_dict()

        sku_id = new_id()
        p_code = payload.product_code or f"SKU-{sku_id[:6].upper()}"
        b_code = payload.barcode or f"890{sku_id[:8]}"
        sku = SKU(
            id=sku_id,
            product_code=p_code,
            barcode=b_code,
            name=payload.name.strip(),
            category=payload.category or "Retail",
            brand=payload.brand or "",
            vendor_id=payload.vendor_id,
            vendor_name=payload.vendor_name or "",
            unit=payload.unit or "Piece",
            unit_cost=payload.unit_cost,
            mrp=payload.mrp or (payload.unit_cost * 1.5),
            selling_price=payload.selling_price or payload.unit_cost,
            unit_price=payload.selling_price or payload.unit_cost,
            min_stock=payload.min_stock,
            reorder_level=payload.reorder_level,
            store_qty=payload.store_qty,
            floor_qty=payload.floor_qty,
            retail_qty=payload.retail_qty,
            created_at=now_utc(),
            updated_at=now_utc(),
        )
        session.add(sku)
        if payload.store_qty > 0:
            session.add(SKUBatch(
                id=new_id(), sku_id=sku_id, qty=payload.store_qty, location="store",
                unit_cost=payload.unit_cost, received_at=now_utc(),
            ))
        if payload.floor_qty > 0:
            session.add(SKUBatch(
                id=new_id(), sku_id=sku_id, qty=payload.floor_qty, location="floor",
                unit_cost=payload.unit_cost, received_at=now_utc(),
            ))
        if payload.retail_qty > 0:
            session.add(SKUBatch(
                id=new_id(), sku_id=sku_id, qty=payload.retail_qty, location="retail",
                unit_cost=payload.unit_cost, received_at=now_utc(),
            ))
        await session.commit()
        return sku.to_dict()

@api.put("/inventory/product/{sku_id}")
async def update_inventory_product(sku_id: str, payload: InventoryProductIn, user: dict = Depends(require_role("owner", "admin", "manager"))):
    async with async_session() as session:
        result = await session.execute(select(SKU).where(SKU.id == sku_id))
        sku = result.scalar_one_or_none()
        if not sku:
            raise HTTPException(404, "Product not found")
        sku.name = payload.name.strip()
        sku.category = payload.category or sku.category
        sku.brand = payload.brand or sku.brand
        sku.vendor_name = payload.vendor_name or sku.vendor_name
        sku.unit = payload.unit or sku.unit
        sku.unit_cost = payload.unit_cost
        sku.mrp = payload.mrp
        sku.selling_price = payload.selling_price
        sku.min_stock = payload.min_stock
        sku.reorder_level = payload.reorder_level
        sku.store_qty = payload.store_qty if payload.store_qty is not None else sku.store_qty
        sku.updated_at = now_utc()
        await session.commit()
        return sku.to_dict()

@api.delete("/inventory/product/{sku_id}")
async def delete_inventory_product(sku_id: str, user: dict = Depends(require_role("owner", "admin", "manager"))):
    async with async_session() as session:
        result = await session.execute(select(SKU).where(SKU.id == sku_id))
        sku = result.scalar_one_or_none()
        if not sku:
            raise HTTPException(404, "Product not found")
        p_name = sku.name

        # Cleanly delete associated batches and ledger logs first to prevent FK failures
        await session.execute(delete(SKUBatch).where(SKUBatch.sku_id == sku_id))
        await session.execute(delete(StockLedger).where(StockLedger.sku_id == sku_id))
        await session.delete(sku)
        await session.commit()
        return {"status": "success", "message": f"Successfully deleted product '{p_name}'"}

@api.get("/inventory/product/{sku_id}/details")
async def get_product_details(sku_id: str):
    async with async_session() as session:
        result = await session.execute(select(SKU).options(selectinload(SKU.batches)).where(SKU.id == sku_id))
        sku = result.scalar_one_or_none()
        if not sku:
            raise HTTPException(404, "Product not found")
        
        # Ledger movements timeline
        ledger_q = select(StockLedger).where(StockLedger.sku_id == sku_id).order_by(StockLedger.timestamp.desc()).limit(50)
        ledger_res = await session.execute(ledger_q)
        timeline = [l.to_dict() for l in ledger_res.scalars().all()]

        # Purchase invoices for history
        pi_q = select(PurchaseInvoiceLine).where(PurchaseInvoiceLine.sku_id == sku_id).limit(20)
        pi_res = await session.execute(pi_q)
        purchase_history = [p.to_dict() for p in pi_res.scalars().all()]

        doc = sku.to_dict()
        doc["valuation"] = round(doc["current_stock"] * (sku.unit_cost or 0), 2)
        doc["timeline"] = timeline
        doc["purchase_history"] = purchase_history
        return doc

@api.post("/inventory/purchase-in")
async def process_purchase_in(payload: PurchaseInPayload, user: dict = Depends(require_role("owner", "admin", "manager"))):
    if not payload.items:
        raise HTTPException(400, "Purchase in items list cannot be empty")
    
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    time_str = datetime.now(timezone.utc).strftime("%H:%M:%S")
    now_iso = now_utc()
    total_invoice_cost = 0.0

    async with async_session() as session:
        for item in payload.items:
            sku_id = item.get("sku_id")
            qty = float(item.get("quantity", 0))
            if qty <= 0:
                continue
            
            result = await session.execute(select(SKU).where(SKU.id == sku_id))
            sku = result.scalar_one_or_none()
            if not sku:
                continue

            before_stock = (sku.store_qty or 0) + (sku.floor_qty or 0) + (sku.retail_qty or 0)
            sku.store_qty = (sku.store_qty or 0) + qty
            sku.unit_cost = float(item.get("purchase_price", sku.unit_cost or 0))
            if item.get("mrp"):
                sku.mrp = float(item["mrp"])
            sku.updated_at = now_iso
            after_stock = before_stock + qty

            line_cost = qty * sku.unit_cost
            total_invoice_cost += line_cost

            # Batch Entry
            session.add(SKUBatch(
                id=new_id(),
                sku_id=sku.id,
                qty=qty,
                location="store",
                expiry_date=item.get("expiry_date"),
                unit_cost=sku.unit_cost,
                batch_number=item.get("batch_number", f"BATCH-{now_iso[:10]}"),
                received_at=now_iso,
            ))

            # Immutable Stock Ledger Entry
            session.add(StockLedger(
                id=new_id(),
                transaction_id=payload.invoice_number or payload.po_number or f"PIN-{new_id()[:8]}",
                date=today_str,
                time=time_str,
                sku_id=sku.id,
                product_code=sku.product_code or f"SKU-{sku.id[:6].upper()}",
                product_name=sku.name,
                store="Main Salon Store",
                transaction_type="purchase_order" if payload.method == "po" else "direct_purchase",
                quantity=qty,
                before_stock=before_stock,
                after_stock=after_stock,
                performed_by=user.get("name") or user.get("email", "Staff"),
                approved_by="System Verified",
                remarks=payload.remarks or f"Invoice #{payload.invoice_number} from {payload.vendor_name}",
                timestamp=now_iso,
            ))

        await session.commit()
    return {"status": "success", "invoice_number": payload.invoice_number, "total_cost": round(total_invoice_cost, 2)}

@api.post("/inventory/stock-out/manual")
async def process_manual_stock_out(payload: ManualStockOutPayload, user: dict = Depends(require_role("owner", "admin", "manager"))):
    if payload.quantity <= 0:
        raise HTTPException(400, "Quantity must be greater than 0")

    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    time_str = datetime.now(timezone.utc).strftime("%H:%M:%S")
    now_iso = now_utc()

    async with async_session() as session:
        result = await session.execute(select(SKU).where(SKU.id == payload.sku_id))
        sku = result.scalar_one_or_none()
        if not sku:
            raise HTTPException(404, "Product not found")

        before_stock = (sku.store_qty or 0) + (sku.floor_qty or 0) + (sku.retail_qty or 0)
        if before_stock < payload.quantity:
            raise HTTPException(400, f"Insufficient stock! Available: {before_stock} {sku.unit}, requested: {payload.quantity}")

        # Deduct from store_qty first, then floor_qty
        rem = payload.quantity
        if (sku.store_qty or 0) >= rem:
            sku.store_qty -= rem
        else:
            rem -= (sku.store_qty or 0)
            sku.store_qty = 0
            sku.floor_qty = max(0, (sku.floor_qty or 0) - rem)

        after_stock = (sku.store_qty or 0) + (sku.floor_qty or 0) + (sku.retail_qty or 0)
        sku.updated_at = now_iso

        # Immutable Stock Ledger Entry
        ledger = StockLedger(
            id=new_id(),
            transaction_id=f"SO-MANUAL-{new_id()[:8]}",
            date=today_str,
            time=time_str,
            sku_id=sku.id,
            product_code=sku.product_code or f"SKU-{sku.id[:6].upper()}",
            product_name=sku.name,
            store="Main Salon Store",
            transaction_type="manual_stock_out",
            quantity=-payload.quantity,
            before_stock=before_stock,
            after_stock=after_stock,
            performed_by=user.get("name") or user.get("email", "Staff"),
            approved_by=payload.issued_to,
            remarks=f"Reason: {payload.reason} | {payload.remarks or ''}",
            timestamp=now_iso,
        )
        session.add(ledger)
        await session.commit()
        return ledger.to_dict()

@api.post("/inventory/stock-out/pos-service")
async def process_pos_service_stock_out(payload: POSServiceStockOutPayload, user: dict = Depends(require_role("owner", "admin", "manager"))):
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    time_str = datetime.now(timezone.utc).strftime("%H:%M:%S")
    now_iso = now_utc()

    async with async_session() as session:
        # Find matching recipe
        res = await session.execute(
            select(ServiceRecipe).options(selectinload(ServiceRecipe.ingredients)).where(ServiceRecipe.service_name == payload.service_name)
        )
        recipe = res.scalar_one_or_none()
        if not recipe or not recipe.ingredients:
            return {"status": "unmapped", "message": f"No Recipe (BOM) mapped for '{payload.service_name}'"}

        deductions = []
        for ing in recipe.ingredients:
            sku_res = await session.execute(select(SKU).where(SKU.id == ing.sku_id))
            sku = sku_res.scalar_one_or_none()
            if not sku:
                continue
            before_stock = (sku.store_qty or 0) + (sku.floor_qty or 0) + (sku.retail_qty or 0)
            deduct_qty = ing.quantity_per_service
            
            # Deduct from floor_qty first for technical service ingredients
            if (sku.floor_qty or 0) >= deduct_qty:
                sku.floor_qty -= deduct_qty
            else:
                rem = deduct_qty - (sku.floor_qty or 0)
                sku.floor_qty = 0
                sku.store_qty = max(0, (sku.store_qty or 0) - rem)

            after_stock = (sku.store_qty or 0) + (sku.floor_qty or 0) + (sku.retail_qty or 0)
            sku.updated_at = now_iso

            session.add(StockLedger(
                id=new_id(),
                transaction_id=payload.pos_transaction_id or f"POS-BOM-{new_id()[:8]}",
                date=today_str,
                time=time_str,
                sku_id=sku.id,
                product_code=sku.product_code or f"SKU-{sku.id[:6].upper()}",
                product_name=sku.name,
                store="Service Floor",
                transaction_type="pos_service_consumption",
                quantity=-deduct_qty,
                before_stock=before_stock,
                after_stock=after_stock,
                performed_by=payload.performed_by or user.get("name", "Stylist"),
                approved_by="POS System Auto",
                remarks=f"BOM Consumption for Service: {payload.service_name}",
                timestamp=now_iso,
            ))
            deductions.append({"sku_name": sku.name, "quantity_deducted": deduct_qty, "unit": ing.unit})

        await session.commit()
    return {"status": "success", "service_name": payload.service_name, "deductions": deductions}

@api.get("/inventory/ledger")
async def get_stock_ledger(
    sku_id: Optional[str] = None,
    transaction_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 200,
):
    async with async_session() as session:
        q = select(StockLedger).order_by(StockLedger.timestamp.desc()).limit(limit)
        if sku_id:
            q = q.where(StockLedger.sku_id == sku_id)
        if transaction_type and transaction_type.lower() != "all":
            q = q.where(StockLedger.transaction_type == transaction_type)
        if date_from:
            q = q.where(StockLedger.date >= date_from)
        if date_to:
            q = q.where(StockLedger.date <= date_to)

        res = await session.execute(q)
        return [l.to_dict() for l in res.scalars().all()]

@api.get("/inventory/dashboard-kpis")
async def get_inventory_kpis():
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    async with async_session() as session:
        skus_res = await session.execute(select(SKU))
        skus = skus_res.scalars().all()

        total_products = len(skus)
        total_value = sum(((s.store_qty or 0) + (s.floor_qty or 0) + (s.retail_qty or 0)) * (s.unit_cost or 0) for s in skus)
        
        low_stock_count = 0
        out_of_stock_count = 0
        category_map = defaultdict(lambda: {"count": 0, "value": 0, "stock": 0})

        for s in skus:
            curr = (s.store_qty or 0) + (s.floor_qty or 0) + (s.retail_qty or 0)
            if curr == 0:
                out_of_stock_count += 1
            elif curr <= (s.min_stock or 5):
                low_stock_count += 1

            cat = s.category or "Others"
            category_map[cat]["count"] += 1
            category_map[cat]["value"] += curr * (s.unit_cost or 0)
            category_map[cat]["stock"] += curr

        # Today stock movement
        today_ledger_res = await session.execute(select(StockLedger).where(StockLedger.date == today_str))
        today_logs = today_ledger_res.scalars().all()

        today_stock_in = sum(l.quantity for l in today_logs if l.quantity > 0)
        today_stock_out = sum(abs(l.quantity) for l in today_logs if l.quantity < 0)

        category_breakdown = [
            {"category": k, "product_count": v["count"], "valuation": round(v["value"], 2), "total_stock": v["stock"]}
            for k, v in category_map.items()
        ]

        return {
            "total_products": total_products,
            "total_inventory_value": round(total_value, 2),
            "low_stock_count": low_stock_count,
            "out_of_stock_count": out_of_stock_count,
            "today_stock_in": today_stock_in,
            "today_stock_out": today_stock_out,
            "category_breakdown": category_breakdown,
        }

# ══════════════════════════════════════════════════════════════
#  EXISTING MODULES — Analytics, Attendance, Audit, COGS, Budget,
#                Procurement, Vendor Management
# ══════════════════════════════════════════════════════════════

# --- Module 2: Sales Analytics ---
@api.get("/analytics/sales")
async def analytics_sales(date_from: Optional[str] = None, date_to: Optional[str] = None):
    async with async_session() as session:
        q = select(
            POSTransaction.type,
            func.sum(POSTransaction.net_price).label("revenue"),
            func.count().label("count"),
            func.sum(POSTransaction.total_discount).label("discounts"),
        ).group_by(POSTransaction.type)
        if date_from: q = q.where(POSTransaction.date >= date_from)
        if date_to: q = q.where(POSTransaction.date <= date_to)
        result = await session.execute(q)
        breakdown = []
        for row in result.all():
            breakdown.append({"type": row[0], "revenue": round(row[1] or 0, 2),
                              "count": row[2], "total_discount": round(row[3] or 0, 2)})
        return {"date_from": date_from, "date_to": date_to, "breakdown": breakdown}

@api.get("/analytics/peak-hours")
async def analytics_peak_hours(date_val: Optional[str] = Query(None, alias="date")):
    async with async_session() as session:
        q = select(POSTransaction.time, func.count(), func.sum(POSTransaction.net_price))
        if date_val: q = q.where(POSTransaction.date == date_val)
        q = q.group_by(POSTransaction.time)
        result = await session.execute(q)
        hours = []
        for row in result.all():
            time_str = row[0] or ""
            hour = time_str[:2] if len(time_str) >= 2 else time_str
            hours.append({"time": time_str, "hour": hour, "count": row[1], "revenue": round(row[2] or 0, 2)})
        return {"date": date_val, "hours": sorted(hours, key=lambda x: x["time"])}

@api.get("/analytics/staff-performance")
async def analytics_staff_performance(month: Optional[str] = None):
    async with async_session() as session:
        q = (
            select(
                POSTransactionStaff.name,
                POSTransaction.type,
                func.sum(POSTransaction.net_price * POSTransactionStaff.pct / 100).label("revenue"),
                func.count().label("count"),
            )
            .join(POSTransaction)
            .group_by(POSTransactionStaff.name, POSTransaction.type)
        )
        if month: q = q.where(POSTransaction.date.like(f"{month}%"))
        result = await session.execute(q)
        staff_data = defaultdict(lambda: {"service_revenue": 0, "retail_revenue": 0, "service_count": 0, "retail_count": 0})
        for row in result.all():
            key = row[0]
            if str(row[1]).lower() == "service":
                staff_data[key]["service_revenue"] = round(row[2] or 0, 2)
                staff_data[key]["service_count"] = row[3]
            elif str(row[1]).lower() == "product":
                staff_data[key]["retail_revenue"] = round(row[2] or 0, 2)
                staff_data[key]["retail_count"] = row[3]
        return {"month": month, "staff": [{"name": k, **v} for k, v in staff_data.items()]}

# --- Module 3: Attendance ---
@api.post("/attendance/clock-in")
async def clock_in(staff_id: str = Query(...), user: dict = Depends(get_current_user)):
    async with async_session() as session:
        staff = (await session.execute(select(Staff).where(Staff.id == staff_id))).scalar_one_or_none()
        if not staff: raise HTTPException(404, "Staff not found")
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        existing = (await session.execute(
            select(Attendance).where(Attendance.staff_id == staff_id, Attendance.date == today)
        )).scalar_one_or_none()
        now = now_utc()
        if existing and existing.clock_in:
            raise HTTPException(400, "Already clocked in today")
        if existing:
            existing.clock_in = now
            existing.status = "present"
        else:
            session.add(Attendance(
                id=new_id(), staff_id=staff_id, staff_name=staff.name,
                date=today, clock_in=now, status="present", created_at=now,
            ))
        await session.commit()
    return {"ok": True, "staff_name": staff.name, "clock_in": now, "date": today}

@api.post("/attendance/clock-out")
async def clock_out(staff_id: str = Query(...), user: dict = Depends(get_current_user)):
    async with async_session() as session:
        staff = (await session.execute(select(Staff).where(Staff.id == staff_id))).scalar_one_or_none()
        if not staff: raise HTTPException(404, "Staff not found")
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        record = (await session.execute(
            select(Attendance).where(Attendance.staff_id == staff_id, Attendance.date == today)
        )).scalar_one_or_none()
        if not record or not record.clock_in: raise HTTPException(400, "No clock-in found")
        if record.clock_out: raise HTTPException(400, "Already clocked out")
        now = now_utc()
        hours = (datetime.fromisoformat(now) - datetime.fromisoformat(record.clock_in)).total_seconds() / 3600
        record.clock_out = now
        record.hours_worked = round(hours, 2)
        record.overtime_hours = round(max(0, hours - 9.0), 2)
        await session.commit()
    return {"ok": True, "staff_name": staff.name, "hours_worked": round(hours, 2)}

@api.get("/attendance")
async def list_attendance(date_val: Optional[str] = Query(None, alias="date"),
                          staff_id: Optional[str] = None, month: Optional[str] = None):
    async with async_session() as session:
        q = select(Attendance)
        if date_val: q = q.where(Attendance.date == date_val)
        if staff_id: q = q.where(Attendance.staff_id == staff_id)
        if month: q = q.where(Attendance.date.like(f"{month}%"))
        q = q.order_by(Attendance.date.desc()).limit(500)
        result = await session.execute(q)
        return [a.to_dict() for a in result.scalars().all()]

@api.get("/attendance/summary")
async def attendance_summary(month: str):
    async with async_session() as session:
        staff_list = (await session.execute(select(Staff))).scalars().all()
        summaries = []
        for s in staff_list:
            q = (
                select(Attendance.status, func.count(), func.sum(Attendance.hours_worked), func.sum(Attendance.overtime_hours))
                .where(Attendance.staff_id == s.id, Attendance.date.like(f"{month}%"))
                .group_by(Attendance.status)
            )
            result = await session.execute(q)
            present = absent = half_day = leave = 0
            total_hours = total_overtime = 0.0
            for row in result.all():
                if row[0] == "present": present = row[1]; total_hours += row[2] or 0; total_overtime += row[3] or 0
                elif row[0] == "absent": absent = row[1]
                elif row[0] == "half_day": half_day = row[1]; total_hours += row[2] or 0
                elif row[0] == "leave": leave = row[1]
            effective_days = present + half_day * 0.5
            calculated_salary = round(s.base_salary / 30 * effective_days, 2)
            summaries.append({
                "staff_id": s.id, "staff_name": s.name, "base_salary": s.base_salary,
                "days_present": present, "days_absent": absent, "days_half_day": half_day, "days_leave": leave,
                "total_hours": round(total_hours, 2), "total_overtime": round(total_overtime, 2),
                "effective_days": effective_days, "calculated_salary": calculated_salary,
            })
        return {"month": month, "summaries": summaries}

# --- Module 7: Stock Audit ---
@api.post("/audit/start")
async def audit_start(user: dict = Depends(require_role("owner", "admin", "manager"))):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    audit_id = new_id()
    async with async_session() as session:
        session.add(StockAudit(
            id=audit_id, audit_date=today, audited_by=user["email"],
            status="in_progress", created_at=now_utc(),
        ))
        await session.commit()
    return {"ok": True, "audit_id": audit_id, "audit_date": today}

@api.put("/audit/{audit_id}/items")
async def audit_submit_items(audit_id: str, items: list, user: dict = Depends(require_role("owner", "admin", "manager"))):
    async with async_session() as session:
        audit = (await session.execute(select(StockAudit).where(StockAudit.id == audit_id))).scalar_one_or_none()
        if not audit: raise HTTPException(404, "Audit not found")
        total_variance = 0.0
        total_value = 0.0
        breakdown = defaultdict(int)
        for item in items:
            sku = (await session.execute(select(SKU).where(SKU.id == item["sku_id"]))).scalar_one_or_none()
            if not sku: continue
            expected = (sku.store_qty or 0) + (sku.floor_qty or 0) + (sku.retail_qty or 0)
            actual = float(item.get("actual_qty", 0))
            variance = actual - expected
            variance_pct = round((variance / expected * 100) if expected else 0, 2)
            reason = item.get("discrepancy_reason")
            if reason: breakdown[reason] += 1
            total_variance += variance
            total_value += variance * (sku.unit_cost or 0)
            session.add(StockAuditItem(
                id=new_id(), audit_id=audit_id, sku_id=sku.id, sku_name=sku.name,
                expected_qty=expected, actual_qty=actual, variance=variance,
                variance_pct=variance_pct, discrepancy_reason=reason,
                notes=item.get("notes", ""),
            ))
        audit.total_skus_audited = len(items)
        audit.total_variance_units = round(total_variance, 2)
        audit.total_variance_value = round(total_value, 2)
        audit.discrepancy_breakdown = dict(breakdown)
        await session.commit()
    return {"ok": True, "total_variance_units": round(total_variance, 2)}

@api.post("/audit/{audit_id}/complete")
async def audit_complete(audit_id: str, user: dict = Depends(require_role("owner", "admin"))):
    async with async_session() as session:
        audit = (await session.execute(
            select(StockAudit).options(selectinload(StockAudit.items)).where(StockAudit.id == audit_id)
        )).scalar_one_or_none()
        if not audit: raise HTTPException(404, "Audit not found")
        audit.status = "completed"
        # Adjust digital inventory based on actual counts
        for item in audit.items:
            sku = (await session.execute(select(SKU).where(SKU.id == item.sku_id))).scalar_one_or_none()
            if sku and item.variance != 0:
                total_current = (sku.store_qty or 0) + (sku.floor_qty or 0) + (sku.retail_qty or 0)
                if total_current > 0:
                    ratio = item.actual_qty / total_current if total_current else 1
                    sku.store_qty = round((sku.store_qty or 0) * ratio, 2)
                    sku.floor_qty = round((sku.floor_qty or 0) * ratio, 2)
                    sku.retail_qty = round((sku.retail_qty or 0) * ratio, 2)
        await session.commit()
    return {"ok": True, "status": "completed"}

@api.get("/audit/history")
async def audit_history():
    async with async_session() as session:
        q = select(StockAudit).options(selectinload(StockAudit.items)).order_by(StockAudit.audit_date.desc()).limit(50)
        result = await session.execute(q)
        return [a.to_dict() for a in result.scalars().unique().all()]

# --- Module 8: COGS ---
@api.post("/cogs/recipes")
async def create_recipe(payload: dict, user: dict = Depends(require_role("owner", "admin"))):
    async with async_session() as session:
        existing = (await session.execute(
            select(ServiceRecipe).where(ServiceRecipe.service_name == payload["service_name"])
        )).scalar_one_or_none()
        if existing:
            # Delete old ingredients and replace
            await session.execute(delete(RecipeIngredient).where(RecipeIngredient.recipe_id == existing.id))
            recipe = existing
            recipe.category = payload.get("category", "")
            recipe.updated_at = now_utc()
        else:
            recipe = ServiceRecipe(
                id=new_id(), service_name=payload["service_name"],
                category=payload.get("category", ""), created_at=now_utc(), updated_at=now_utc(),
            )
            session.add(recipe)
            await session.flush()
        total_cost = 0.0
        for ing in payload.get("ingredients", []):
            sku = (await session.execute(select(SKU).where(SKU.id == ing["sku_id"]))).scalar_one_or_none()
            cost_per_unit = sku.unit_cost if sku else 0
            ingredient_cost = cost_per_unit * ing["quantity_per_service"]
            total_cost += ingredient_cost
            session.add(RecipeIngredient(
                id=new_id(), recipe_id=recipe.id, sku_id=ing["sku_id"],
                sku_name=sku.name if sku else "", quantity_per_service=ing["quantity_per_service"],
                unit=ing.get("unit", "ml"), cost_per_unit=cost_per_unit,
            ))
        recipe.total_material_cost = round(total_cost, 2)
        await session.commit()
    return {"ok": True, "recipe_id": recipe.id, "total_material_cost": recipe.total_material_cost}

@api.get("/cogs/recipes")
async def list_recipes():
    async with async_session() as session:
        q = select(ServiceRecipe).options(selectinload(ServiceRecipe.ingredients))
        result = await session.execute(q)
        return [r.to_dict() for r in result.scalars().unique().all()]

@api.get("/cogs/analysis")
async def cogs_analysis(month: Optional[str] = None):
    async with async_session() as session:
        q = select(ServiceConsumptionLog)
        if month: q = q.where(ServiceConsumptionLog.date.like(f"{month}%"))
        result = await session.execute(q)
        logs = result.scalars().all()
        by_service = defaultdict(lambda: {"total_cost": 0, "total_revenue": 0, "count": 0})
        for log_entry in logs:
            key = log_entry.service_name
            by_service[key]["total_cost"] += log_entry.total_cost
            by_service[key]["total_revenue"] += log_entry.revenue_generated
            by_service[key]["count"] += 1
        analysis = []
        for service_name, data in by_service.items():
            margin = ((data["total_revenue"] - data["total_cost"]) / data["total_revenue"] * 100) if data["total_revenue"] else 0
            analysis.append({
                "service_name": service_name, "total_cost": round(data["total_cost"], 2),
                "total_revenue": round(data["total_revenue"], 2), "count": data["count"],
                "margin_pct": round(margin, 2),
            })
        return {"month": month, "analysis": analysis}

# --- Module 9: Budgets ---
@api.post("/budgets")
async def create_budget(payload: dict, user: dict = Depends(require_role("owner", "admin"))):
    async with async_session() as session:
        budget_id = new_id()
        budget = Budget(
            id=budget_id, month=payload["month"], category=payload.get("category", "all"),
            budgeted_amount=payload.get("budgeted_amount", 0),
            budgeted_units=payload.get("budgeted_units", 0),
            created_by=user["email"], created_at=now_utc(),
        )
        session.add(budget)
        for item in payload.get("line_items", []):
            sku = (await session.execute(select(SKU).where(SKU.id == item["sku_id"]))).scalar_one_or_none()
            session.add(BudgetLineItem(
                id=new_id(), budget_id=budget_id, sku_id=item["sku_id"],
                sku_name=sku.name if sku else "", budgeted_qty=item.get("budgeted_qty", 0),
                budgeted_cost=item.get("budgeted_cost", 0),
            ))
        await session.commit()
    return {"ok": True, "budget_id": budget_id}

@api.get("/budgets")
async def list_budgets(month: Optional[str] = None):
    async with async_session() as session:
        q = select(Budget).options(selectinload(Budget.line_items))
        if month: q = q.where(Budget.month == month)
        result = await session.execute(q)
        return [b.to_dict() for b in result.scalars().unique().all()]

@api.get("/budgets/variance")
async def budget_variance(month: str):
    async with async_session() as session:
        budgets = (await session.execute(
            select(Budget).options(selectinload(Budget.line_items)).where(Budget.month == month)
        )).scalars().unique().all()
        # Actual consumption from checkouts this month
        actual_q = (
            select(Checkout.sku_id, func.sum(Checkout.quantity), func.sum(Checkout.quantity * SKU.unit_cost))
            .join(SKU, Checkout.sku_id == SKU.id)
            .where(Checkout.checked_out_at.like(f"{month}%"))
            .group_by(Checkout.sku_id)
        )
        actual_result = await session.execute(actual_q)
        actual_map = {r[0]: {"qty": r[1], "cost": round(r[2] or 0, 2)} for r in actual_result.all()}
        variance_items = []
        for b in budgets:
            for item in b.line_items:
                actual = actual_map.get(item.sku_id, {"qty": 0, "cost": 0})
                variance_items.append({
                    "sku_id": item.sku_id, "sku_name": item.sku_name,
                    "budgeted_qty": item.budgeted_qty, "actual_qty": actual["qty"],
                    "budgeted_cost": item.budgeted_cost, "actual_cost": actual["cost"],
                    "variance_qty": actual["qty"] - item.budgeted_qty,
                    "variance_cost": round(actual["cost"] - item.budgeted_cost, 2),
                })
        return {"month": month, "variance": variance_items}

# --- Module 10: Procurement PO Lifecycle ---
@api.post("/procurement/purchase-orders")
async def create_po(payload: dict, user: dict = Depends(require_role("owner", "admin"))):
    async with async_session() as session:
        # Auto-generate PO number
        count_result = await session.execute(select(func.count()).select_from(PurchaseOrder))
        count = (count_result.scalar() or 0) + 1
        po_number = f"PO-{datetime.now().strftime('%Y')}-{count:04d}"
        vendor = (await session.execute(select(Vendor).where(Vendor.id == payload["vendor_id"]))).scalar_one_or_none()
        if not vendor: raise HTTPException(404, "Vendor not found")
        po_id = new_id()
        total = 0.0
        po = PurchaseOrder(
            id=po_id, po_number=po_number, vendor_id=vendor.id, vendor_name=vendor.name,
            status="draft", expected_delivery=payload.get("expected_delivery"),
            notes=payload.get("notes", ""), created_by=user["email"], created_at=now_utc(),
        )
        session.add(po)
        for line in payload.get("lines", []):
            sku = (await session.execute(select(SKU).where(SKU.id == line["sku_id"]))).scalar_one_or_none()
            line_total = line["quantity"] * line["unit_cost"]
            total += line_total
            session.add(POLine(
                id=new_id(), po_id=po_id, sku_id=line["sku_id"],
                sku_name=sku.name if sku else "", quantity=line["quantity"],
                unit_cost=line["unit_cost"], line_total=line_total,
            ))
        po.total = round(total, 2)
        session.add(POStatusHistory(
            id=new_id(), po_id=po_id, status="draft", changed_by=user["email"], timestamp=now_utc(),
        ))
        await session.commit()
    return {"ok": True, "po_id": po_id, "po_number": po_number}

@api.put("/procurement/purchase-orders/{po_id}/status")
async def update_po_status(po_id: str, payload: dict, user: dict = Depends(require_role("owner", "admin", "manager"))):
    valid_transitions = {
        "draft": ["requested", "cancelled"],
        "requested": ["ordered", "cancelled"],
        "ordered": ["in_transit", "cancelled"],
        "in_transit": ["delivered", "cancelled"],
        "delivered": ["stocked"],
    }
    async with async_session() as session:
        po = (await session.execute(select(PurchaseOrder).where(PurchaseOrder.id == po_id))).scalar_one_or_none()
        if not po: raise HTTPException(404, "PO not found")
        new_status = payload["status"]
        if new_status not in valid_transitions.get(po.status, []):
            raise HTTPException(400, f"Cannot transition from {po.status} to {new_status}")
        po.status = new_status
        if new_status == "delivered":
            po.actual_delivery = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        session.add(POStatusHistory(
            id=new_id(), po_id=po_id, status=new_status,
            changed_by=user["email"], notes=payload.get("notes", ""), timestamp=now_utc(),
        ))
        # If stocked, add inventory batches
        if new_status == "stocked":
            lines = (await session.execute(select(POLine).where(POLine.po_id == po_id))).scalars().all()
            for line in lines:
                sku = (await session.execute(select(SKU).where(SKU.id == line.sku_id))).scalar_one_or_none()
                if sku:
                    session.add(SKUBatch(
                        id=new_id(), sku_id=line.sku_id, qty=line.quantity,
                        location="store", unit_cost=line.unit_cost, received_at=now_utc(),
                    ))
                    sku.store_qty = (sku.store_qty or 0) + line.quantity
        await session.commit()
    return {"ok": True, "status": new_status}

@api.get("/procurement/purchase-orders")
async def list_pos_orders(status: Optional[str] = None):
    async with async_session() as session:
        q = select(PurchaseOrder).options(
            selectinload(PurchaseOrder.lines), selectinload(PurchaseOrder.status_history)
        )
        if status: q = q.where(PurchaseOrder.status == status)
        q = q.order_by(PurchaseOrder.created_at.desc()).limit(100)
        result = await session.execute(q)
        return [po.to_dict() for po in result.scalars().unique().all()]

# --- Module 11: Vendor Management ---
@api.put("/vendors/{vendor_id}")
async def update_vendor(vendor_id: str, payload: dict, user: dict = Depends(require_role("owner", "admin"))):
    async with async_session() as session:
        vendor = (await session.execute(select(Vendor).where(Vendor.id == vendor_id))).scalar_one_or_none()
        if not vendor: raise HTTPException(404, "Vendor not found")
        for k, v in payload.items():
            if hasattr(vendor, k): setattr(vendor, k, v)
        await session.commit()
        return vendor.to_dict()

@api.get("/vendors/{vendor_id}/scorecard")
async def vendor_scorecard(vendor_id: str):
    async with async_session() as session:
        vendor = (await session.execute(select(Vendor).where(Vendor.id == vendor_id))).scalar_one_or_none()
        if not vendor: raise HTTPException(404, "Vendor not found")
        # Count POs
        po_count = (await session.execute(
            select(func.count()).select_from(PurchaseOrder).where(PurchaseOrder.vendor_id == vendor_id)
        )).scalar() or 0
        on_time = (await session.execute(
            select(func.count()).select_from(PurchaseOrder).where(
                PurchaseOrder.vendor_id == vendor_id,
                PurchaseOrder.status == "stocked",
                PurchaseOrder.actual_delivery <= PurchaseOrder.expected_delivery,
            )
        )).scalar() or 0
        contracts = (await session.execute(
            select(VendorContract).where(VendorContract.vendor_id == vendor_id)
        )).scalars().all()
        return {
            "vendor": vendor.to_dict(),
            "total_orders": po_count,
            "on_time_deliveries": on_time,
            "reliability_pct": round(on_time / po_count * 100, 1) if po_count else 0,
            "contracts": [c.to_dict() for c in contracts],
        }

@api.get("/vendors/matrix")
async def vendor_matrix():
    async with async_session() as session:
        vendors = (await session.execute(select(Vendor).order_by(Vendor.name))).scalars().all()
        matrix = []
        for v in vendors:
            po_count = (await session.execute(
                select(func.count()).select_from(PurchaseOrder).where(PurchaseOrder.vendor_id == v.id)
            )).scalar() or 0
            sku_count = (await session.execute(
                select(func.count()).select_from(SKU).where(SKU.vendor_id == v.id)
            )).scalar() or 0
            matrix.append({
                **v.to_dict(),
                "total_pos": po_count,
                "sku_count": sku_count,
            })
        return matrix

# --- Vendor Contracts ---
@api.post("/vendors/{vendor_id}/contracts")
async def create_vendor_contract(vendor_id: str, payload: dict, user: dict = Depends(require_role("owner", "admin"))):
    async with async_session() as session:
        vendor = (await session.execute(select(Vendor).where(Vendor.id == vendor_id))).scalar_one_or_none()
        if not vendor: raise HTTPException(404, "Vendor not found")
        contract = VendorContract(
            id=new_id(), vendor_id=vendor_id,
            contract_number=payload.get("contract_number", ""),
            start_date=payload["start_date"], end_date=payload["end_date"],
            terms=payload.get("terms", ""),
            discount_structure=payload.get("discount_structure"),
            minimum_order_frequency=payload.get("minimum_order_frequency"),
            sla_delivery_days=payload.get("sla_delivery_days"),
            penalty_clause=payload.get("penalty_clause", ""),
            status="active", created_at=now_utc(),
        )
        session.add(contract)
        await session.commit()
    return {"ok": True, "contract_id": contract.id}

# --- Executive Dashboard ---
@api.get("/dashboard/executive")
async def executive_dashboard():
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    async with async_session() as session:
        # Daily revenue
        daily_rev = (await session.execute(
            select(func.sum(POSTransaction.net_price)).where(POSTransaction.date == today)
        )).scalar() or 0
        # Active staff today
        active_staff = (await session.execute(
            select(func.count()).select_from(Attendance).where(Attendance.date == today, Attendance.status == "present")
        )).scalar() or 0
        # Low stock alerts (below reorder point)
        cfg = await get_config()
        skus = (await session.execute(select(SKU))).scalars().all()
        low_stock = sum(1 for s in skus if ((s.store_qty or 0) + (s.floor_qty or 0) + (s.retail_qty or 0)) < 5)
        # Expiring within 30 days
        cutoff30 = (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%Y-%m-%d")
        expiring_soon = (await session.execute(
            select(func.count()).select_from(SKUBatch).where(
                SKUBatch.expiry_date.isnot(None), SKUBatch.expiry_date <= cutoff30, SKUBatch.qty > 0
            )
        )).scalar() or 0
        # Total staff
        total_staff = (await session.execute(select(func.count()).select_from(Staff))).scalar() or 0
    return {
        "date": today,
        "daily_revenue": round(daily_rev, 2),
        "active_staff_today": active_staff,
        "total_staff": total_staff,
        "low_stock_alerts": low_stock,
        "expiring_batches_30d": expiring_soon,
    }


# ------------------ Startup ------------------
@app.on_event("startup")
async def startup():
    # Create all PostgreSQL tables
    await init_pg()
    log.info(f"PostgreSQL tables created/verified. Known tables: {list(Base.metadata.tables.keys())}")

    try:
        async with async_session() as session:
            # Seed users
            async def seed_user(email_key, pw_key, name, role):
                email = os.environ.get(email_key, "").lower().strip()
                pw = os.environ.get(pw_key, "")
                if not email or not pw:
                    return
                existing = (await session.execute(select(User).where(User.email == email))).scalar_one_or_none()
                if not existing:
                    session.add(User(
                        id=new_id(), email=email, password_hash=hash_pw(pw),
                        name=name, role=role, created_at=now_utc(),
                    ))
                elif not verify_pw(pw, existing.password_hash):
                    existing.password_hash = hash_pw(pw)

            await seed_user("OWNER_EMAIL", "OWNER_PASSWORD", "Salon Owner", "owner")
            await seed_user("MANAGER_EMAIL", "MANAGER_PASSWORD", "Salon Manager", "manager")

            # Seed config
            existing_cfg = (await session.execute(select(AppConfig).where(AppConfig.id == "master"))).scalar_one_or_none()
            if not existing_cfg:
                session.add(AppConfig(id="master", data=DEFAULT_CONFIG))
            else:
                data = dict(existing_cfg.data or {})
                updated = False
                if "product_incentives" not in data:
                    data["product_incentives"] = DEFAULT_CONFIG["product_incentives"]
                    data["retail_commission_pct"] = 0
                    updated = True
                if "prepaid_card_bonuses" not in data:
                    data["prepaid_card_bonuses"] = DEFAULT_CONFIG["prepaid_card_bonuses"]
                    updated = True
                if updated:
                    existing_cfg.data = data

            await session.commit()
    except Exception as seed_err:
        log.warning(f"Startup seed (users/config) failed — server will still run: {seed_err}")

    # Seed sample POS data on first run
    try:
        async with async_session() as session:
            pos_count = (await session.execute(select(func.count()).select_from(POSTransaction))).scalar() or 0
            if pos_count == 0:
                urls_env = os.environ.get("POS_SAMPLE_URLS") or os.environ.get("POS_SAMPLE_URL", "")
                for sample_url in [u.strip() for u in urls_env.split(",") if u.strip()]:
                    try:
                        r = httpreq.get(sample_url, timeout=15)
                        if r.status_code == 200:
                            result = await import_csv_bytes(r.content)
                            log.info(f"Seeded sample POS from {sample_url}: {result}")
                    except Exception as e:
                        log.warning(f"Failed to seed sample POS {sample_url}: {e}")
    except Exception as e:
        log.warning(f"POS seed check failed: {e}")

    # Seed inventory from remote URLs on first run
    try:
        async with async_session() as session:
            sku_count = (await session.execute(select(func.count()).select_from(SKU))).scalar() or 0
            if sku_count == 0:
                seeds = [
                    ("retail", os.environ.get("RETAIL_STOCK_URL", "")),
                    ("technical", os.environ.get("TECHNICAL_STOCK_URL", "")),
                ]
                for ledger, url in seeds:
                    if not url:
                        continue
                    try:
                        r = httpreq.get(url, timeout=30)
                        if r.status_code == 200:
                            res = await _import_ledger(r.content, ledger)
                            log.info(f"Seeded {ledger}: {res}")
                    except Exception as e:
                        log.warning(f"Ledger seed {ledger} failed: {e}")
    except Exception as e:
        log.warning(f"Inventory seed check failed: {e}")

    log.info("LSS backend startup complete (PostgreSQL).")


# ------------------ Wire up ------------------
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown():
    await close_pg()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host="127.0.0.1", port=8000, reload=True)
